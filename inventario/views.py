from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.template.loader import get_template
# from xhtml2pdf import pisa  # Eliminado para migrar a WeasyPrint
from django.shortcuts import render, redirect, get_object_or_404 # <--- Agrega get_object_or_404
from .models import (
    Material, Activo, ReporteRecepcion, DetalleRecepcion, GastoDirecto,
    SalidaMaterial, GuiaTraslado, PresupuestoAnual, 
    SalidaMaterialDetalle, CentroCosto, CorreoAutorizado, RegistroActividad
)
from .forms import ReporteRecepcionForm, DetalleRecepcionForm, SalidaMaterialForm, GuiaTrasladoForm, MaterialForm, SalidaMaterialEditForm, DetalleRecepcionEditForm, ReporteRecepcionEditForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q
from django.utils import timezone

def registrar_actividad(user, accion, detalles=""):
    """
    Registra una acción de auditoría en la base de datos de forma segura.
    """
    try:
        if user and user.is_authenticated:
            RegistroActividad.objects.create(usuario=user, accion=accion, detalles=detalles)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error registrando actividad de auditoría: {e}")
from django.db import transaction
from django.core.exceptions import ValidationError
import json
from decimal import Decimal
import datetime as dt

# Para exportación a Excel
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


from functools import wraps
from django.contrib import messages

def roles_requeridos(groups_allowed):
    """
    Decorador para restringir el acceso a vistas basado en roles (Grupos de Django).
    Ejemplo de uso: @roles_requeridos(['Coordinador de Almacén', 'Especialista de Activos'])
    """
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')
            
            # Superusuario tiene acceso total automático
            if request.user.is_superuser:
                return view_func(request, *args, **kwargs)
                
            # Validar pertenencia a los grupos autorizados
            user_groups = request.user.groups.values_list('name', flat=True)
            is_authorized = any(group in user_groups for group in groups_allowed)
            
            if not is_authorized:
                messages.error(request, "Acceso Denegado: Su rol en el sistema no tiene permisos para realizar esta acción.")
                referer = request.META.get('HTTP_REFERER')
                if referer and not referer.endswith(request.path):
                    return redirect(referer)
                return redirect('dashboard')
                
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator


# Función para saber si el usuario es Operador o Jefe
def es_almacenista(user):
    return user.is_staff or user.is_superuser

# VISTA 1: El Dashboard (Gráficas y Resumen)
@login_required(login_url='login')
def dashboard(request):
    from decimal import Decimal
    import json
    import datetime
    from django.db.models import Sum, Count
    from django.db.models.functions import TruncMonth
    
    total_materiales_count = Material.objects.count()
    # Contamos cuántos materiales tienen stock crítico (menor a 5)
    alertas_stock = Material.objects.filter(stock_actual__lt=5).count()
    # Entradas de hoy (RP)
    entradas_hoy = ReporteRecepcion.objects.filter(fecha_recepcion=timezone.now()).count()
    # Auditoría: Últimas 15 actividades con select_related para evitar query N+1
    actividades = RegistroActividad.objects.select_related('usuario', 'usuario__perfil').order_by('-fecha')[:15]

    # --- NUEVOS KPIS DE VALORIZACIÓN Y AUDITORÍA ---
    from django.core.cache import cache
    
    kpis_val = cache.get('kpis_valorizacion')
    if kpis_val is None:
        # 1. Valorización de Materiales (excluyendo Directo al Gasto)
        materiales_qs = Material.objects.exclude(tipo='DIRECTO AL GASTO').prefetch_related('entradas__detalles_salida')
        total_materiales_val = sum(m.valor_total_inventario for m in materiales_qs)

        # 2. Valorización de Activos
        entradas_activos = DetalleRecepcion.objects.filter(tipo_ingreso='Activo', activo__isnull=False).prefetch_related('detalles_salida')
        total_activos_val = sum(lote.cantidad_disponible * (lote.precio_unitario or Decimal('0.00')) for lote in entradas_activos)

        total_dinero_inventario = total_materiales_val + total_activos_val
        kpis_val = {
            'total_materiales_val': str(total_materiales_val),
            'total_activos_val': str(total_activos_val),
            'total_dinero_inventario': str(total_dinero_inventario),
        }
        cache.set('kpis_valorizacion', kpis_val, 900) # Cache por 15 minutos
    else:
        total_materiales_val = Decimal(kpis_val['total_materiales_val'])
        total_activos_val = Decimal(kpis_val['total_activos_val'])
        total_dinero_inventario = Decimal(kpis_val['total_dinero_inventario'])

    # 3. Últimos movimientos (Recepción/Entradas)
    em_reciente = DetalleRecepcion.objects.filter(nro_control_entrada__startswith='EM').order_by('-id').first()
    ultima_em = em_reciente.nro_control_entrada if em_reciente else "N/A"
    
    ea_reciente = DetalleRecepcion.objects.filter(nro_control_entrada__startswith='EA').order_by('-id').first()
    ultima_ea = ea_reciente.nro_control_entrada if ea_reciente else "N/A"
    
    edg_reciente = DetalleRecepcion.objects.filter(nro_control_entrada__startswith='EDG').order_by('-id').first()
    ultimo_edg = edg_reciente.nro_control_entrada if edg_reciente else "N/A"

    # 4. Últimos movimientos (Salidas/Despachos)
    sm_reciente = SalidaMaterial.objects.filter(tipo_salida='SM').order_by('-id').first()
    ultima_sm = sm_reciente.codigo_salida_completo if sm_reciente else "N/A"
    
    sa_reciente = SalidaMaterial.objects.filter(tipo_salida='SA').order_by('-id').first()
    ultima_sa = sa_reciente.codigo_salida_completo if sa_reciente else "N/A"

    # --- 📊 1. GRÁFICO DE DONA: CONSUMO POR DEPARTAMENTO ---
    consumo_depto = SalidaMaterial.objects.exclude(
        departamento__isnull=True
    ).exclude(
        departamento=''
    ).values('departamento').annotate(
        total_cantidad=Sum('cantidad')
    ).order_by('-total_cantidad')

    labels_departamentos = [item['departamento'] for item in consumo_depto]
    data_departamentos = [float(item['total_cantidad']) for item in consumo_depto]

    # --- 📈 2. GRÁFICO DE LÍNEAS: FLUJO MENSUAL DE MOVIMIENTOS ---
    hoy = timezone.now().date()
    meses_tuplas = []
    a, m = hoy.year, hoy.month
    for _ in range(6):
        meses_tuplas.append((a, m))
        m -= 1
        if m == 0:
            m = 12
            a -= 1
    meses_tuplas.reverse()

    NOMBRES_MESES = {
        1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
        7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
    }
    labels_meses = [NOMBRES_MESES[mes] for _, mes in meses_tuplas]

    # Primer día del mes más antiguo
    start_year, start_month = meses_tuplas[0]
    fecha_inicio = datetime.date(start_year, start_month, 1)

    # Entradas y salidas agrupadas
    entradas_qs = DetalleRecepcion.objects.filter(
        fecha_recepcion__gte=fecha_inicio
    ).annotate(
        mes=TruncMonth('fecha_recepcion')
    ).values('mes').annotate(
        total=Count('id')
    ).order_by('mes')

    salidas_qs = SalidaMaterial.objects.filter(
        fecha_despacho__gte=fecha_inicio
    ).annotate(
        mes=TruncMonth('fecha_despacho')
    ).values('mes').annotate(
        total=Count('id')
    ).order_by('mes')

    entradas_dict = {}
    for e in entradas_qs:
        if e['mes']:
            entradas_dict[(e['mes'].year, e['mes'].month)] = e['total']

    salidas_dict = {}
    for s in salidas_qs:
        if s['mes']:
            salidas_dict[(s['mes'].year, s['mes'].month)] = s['total']

    data_entradas = [entradas_dict.get(key, 0) for key in meses_tuplas]
    data_salidas = [salidas_dict.get(key, 0) for key in meses_tuplas]

    # --- 🎂 3. GRÁFICO DE PASTEL: DISTRIBUCIÓN DEL DINERO ---
    data_dinero = [float(total_activos_val), float(total_materiales_val)]

    contexto = {
        'total_materiales': total_materiales_count,
        'alertas_stock': alertas_stock,
        'entradas_hoy': entradas_hoy,
        'actividades': actividades,
        'total_dinero_inventario': total_dinero_inventario,
        'ultima_em': ultima_em,
        'ultima_ea': ultima_ea,
        'ultimo_edg': ultimo_edg,
        'ultima_sm': ultima_sm,
        'ultima_sa': ultima_sa,
        
        # Serializaciones JSON para el Frontend
        'grafico_departamentos_labels': json.dumps(labels_departamentos),
        'grafico_departamentos_data': json.dumps(data_departamentos),
        'grafico_flujo_labels': json.dumps(labels_meses),
        'grafico_flujo_entradas': json.dumps(data_entradas),
        'grafico_flujo_salidas': json.dumps(data_salidas),
        'grafico_dinero_data': json.dumps(data_dinero),
    }
    return render(request, 'inventario/dashboard.html', contexto)

# VISTA 2: Registro Maestro (La tabla pura)
@login_required(login_url='login')
def lista_materiales(request):
    # 1. Obtener parámetros de filtro por columna
    f_rq = request.GET.get('f_rq', '').strip()
    f_codigo = request.GET.get('f_codigo', '').strip()
    f_desc = request.GET.get('f_desc', '').strip()
    f_np = request.GET.get('f_np', '').strip()
    f_odc = request.GET.get('f_odc', '').strip()
    f_em = request.GET.get('f_em', '').strip()
    f_prov = request.GET.get('f_prov', '').strip()
    f_tipo = request.GET.get('f_tipo', '').strip()
    f_cargo = request.GET.get('f_cargo', '').strip()
    f_nota = request.GET.get('f_nota', '').strip()

    materiales_qs = Material.objects.exclude(tipo='DIRECTO AL GASTO')
    activos_qs = Activo.objects.filter(stock__gt=0)

    # 2. Aplicar filtros condicionales (Server-side)
    if f_rq:
        materiales_qs = materiales_qs.filter(detallerecepcion__nro_rq__icontains=f_rq)
    if f_codigo:
        materiales_qs = materiales_qs.filter(codigo__icontains=f_codigo)
        activos_qs = activos_qs.filter(codigo_activo__icontains=f_codigo)
    if f_desc:
        materiales_qs = materiales_qs.filter(descripcion__icontains=f_desc)
        activos_qs = activos_qs.filter(descripcion__icontains=f_desc)
    if f_np:
        materiales_qs = materiales_qs.filter(nro_parte__icontains=f_np)
    if f_odc:
        materiales_qs = materiales_qs.filter(detallerecepcion__nro_odc__icontains=f_odc)
    if f_em:
        materiales_qs = materiales_qs.filter(detallerecepcion__nro_control_entrada__icontains=f_em)
    if f_prov:
        materiales_qs = materiales_qs.filter(detallerecepcion__proveedor__icontains=f_prov)
    if f_tipo:
        materiales_qs = materiales_qs.filter(tipo__icontains=f_tipo)
    if f_cargo:
        materiales_qs = materiales_qs.filter(cargo__icontains=f_cargo)
    if f_nota:
        materiales_qs = materiales_qs.filter(detallerecepcion__nro_nota_entrega__icontains=f_nota)

    # 3. Optimización y Paginación
    from django.db.models import Prefetch
    materiales_qs = materiales_qs.order_by('codigo').prefetch_related(
        Prefetch('entradas', queryset=DetalleRecepcion.objects.order_by('fecha_recepcion', 'id')),
        'entradas__detalles_salida'
    ).distinct()

    paginator = Paginator(materiales_qs, 50)
    page_number = request.GET.get('page')
    materiales_paginados = paginator.get_page(page_number)

    # Paginación independiente para Activos
    activos_qs = activos_qs.order_by('codigo_activo')
    paginator_activos = Paginator(activos_qs, 50)
    page_activos = request.GET.get('page_activos')
    activos_paginados = paginator_activos.get_page(page_activos)

    # 4. Preservar estado para el HTML y enlaces de página
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    if 'page_activos' in query_params:
        del query_params['page_activos']
    querystring = query_params.urlencode()
    query_prefix = f"{querystring}&" if querystring else ''

    # 5. Departamentos para el reporte de valoración
    departamentos = PresupuestoAnual.objects.values_list('departamento', flat=True).distinct().order_by('departamento')

    contexto = {
        'materiales': materiales_paginados,
        'activos': activos_paginados,
        'query_prefix': query_prefix,
        'filtros': {
            'rq': f_rq, 'codigo': f_codigo, 'desc': f_desc, 'np': f_np,
            'odc': f_odc, 'em': f_em, 'prov': f_prov, 'tipo': f_tipo,
            'cargo': f_cargo, 'nota': f_nota
        },
        'departamentos': departamentos,
    }
    return render(request, 'inventario/lista_materiales.html', contexto)

@login_required(login_url='login')
@roles_requeridos(['Coordinador de Almacén', 'Especialista de Activos'])
def crear_material(request):
    """
    Router de creación para el Registro Maestro con trazabilidad total.
    """
    if request.method == 'POST':
        print("\n[CHECKPOINT 1] Petición POST recibida en crear_material")
        print(f"DEBUG: Datos POST: {request.POST}")

        # Captura del Enrutador
        tipo_item = request.POST.get('tipo_material')
        print(f"[CHECKPOINT 2] Tipo seleccionado: {tipo_item}")
        
        try:
            from django.contrib import messages
            from django.db import transaction
            
            with transaction.atomic():
                # Extracción y Limpieza
                codigo = request.POST.get('codigo', '').strip().upper()
                descripcion = request.POST.get('descripcion', '').strip().upper()
                ubicacion = request.POST.get('ubicacion', '').strip().upper()
                
                print(f"[CHECKPOINT 3] Procesando ítem: {codigo}")

                if not codigo or not descripcion:
                    print("ERROR: Código o Descripción vacíos.")
                    raise ValueError("El Código y la Descripción son campos obligatorios.")

                if not tipo_item:
                    print("ERROR: El campo tipo_material llegó vacío.")
                    raise ValueError("No se recibió el tipo de ítem (Material/Activo). Revise el selector del formulario.")

                if tipo_item == 'ACTIVOS':
                    # --- LÓGICA PARA ACTIVOS FIJOS ---
                    marca = request.POST.get('marca', '').strip().upper()
                    modelo = request.POST.get('modelo', '').strip().upper()
                    serial = request.POST.get('serial', '').strip().upper()
                    
                    # Campos adicionales con sufijo _ac requeridos por el modelo
                    unidad = request.POST.get('unidad_medida', '').strip().upper()
                    cargo_uso = request.POST.get('cargo', 'OPERACIONES')
                    parte = request.POST.get('nro_parte', '').strip().upper()

                    if Activo.objects.filter(codigo_activo=codigo).exists():
                        raise ValueError(f"El código de activo '{codigo}' ya existe en el sistema.")

                    Activo.objects.create(
                        codigo_activo=codigo,
                        descripcion=descripcion,
                        marca=marca,
                        modelo=modelo,
                        serial=serial,
                        stock=0,
                        # Mapeo corregido a las columnas del modelo Activo
                        unidad_medida_ac=unidad,
                        ubicacion=ubicacion,          # Corregido: sin sufijo _ac
                        cargo_ac=cargo_uso,           # Corregido: coincide con models.py
                        nro_parte_ac=parte
                    )
                    registrar_actividad(request.user, "Creó Activo Fijo", f"Activo {codigo}: {descripcion} - Serial: {serial}")
                else:
                    # --- LÓGICA PARA MATERIALES ---
                    unidad_medida = request.POST.get('unidad_medida', '').strip().upper()
                    cargo = request.POST.get('cargo', 'OPERACIONES') # Viene del select dinámico
                    nro_parte = request.POST.get('nro_parte', '').strip().upper()
                    # El campo 'tipo' en el modelo Material (MATERIAL, DIRECTO AL GASTO, etc.)
                    # Usamos el valor capturado en el enrutador
                    tipo_maestro = tipo_item if tipo_item else 'MATERIAL'

                    if Material.objects.filter(codigo=codigo).exists():
                        raise ValueError(f"El código de material '{codigo}' ya existe en el sistema.")

                    Material.objects.create(
                        codigo=codigo,
                        descripcion=descripcion,
                        tipo=tipo_maestro,
                        cargo=cargo,
                        nro_parte=nro_parte,
                        unidad_medida=unidad_medida,
                        ubicacion=ubicacion,
                        stock_actual=0
                    )
                    registrar_actividad(request.user, "Creó Material", f"{tipo_maestro} {codigo}: {descripcion}")

                messages.success(request, f"Éxito: Registro '{codigo}' creado correctamente.")
                return redirect('lista_materiales')

        except Exception as e:
            # 4. Captura de errores silenciosos
            print(f"CRITICAL ERROR EN CREAR_MATERIAL: {str(e)}")
            from django.contrib import messages
            messages.error(request, f"Error al guardar: {str(e)}")
            
    # GET: Preparar formulario y datos dinámicos
    form = MaterialForm()
    departamentos = PresupuestoAnual.objects.values_list('departamento', flat=True).distinct().order_by('departamento')
    
    return render(request, 'inventario/crear_material.html', {
        'form': form,
        'departamentos': departamentos
    })

# Vista 3: Lista de Reportes de Entrada (CON BUSCADOR)
@login_required(login_url='login')
def lista_entradas(request):
    from django.db.models import Q, Sum, F, Max, DecimalField
    # 1. Capturar filtros por columna (TODOS)
    f_base = request.GET.get('f_base', '').strip()
    f_em = request.GET.get('f_em', '').strip()
    f_fecha_rep = request.GET.get('f_fecha_rep', '').strip()
    f_fecha_ent = request.GET.get('f_fecha_ent', '').strip()
    f_odc = request.GET.get('f_odc', '').strip()
    f_nota = request.GET.get('f_nota', '').strip()
    f_prov = request.GET.get('f_prov', '').strip()
    f_mat = request.GET.get('f_mat', '').strip()
    f_obs = request.GET.get('f_obs', '').strip()
    f_rm = request.GET.get('f_rm', '').strip()
    f_vol = request.GET.get('f_vol', '').strip()

    # 2. Obtener los IDs de los registros representativos
    ids_unicos = DetalleRecepcion.objects.exclude(es_saldo_inicial=True).values('nro_control_entrada').annotate(max_id=Max('id')).values_list('max_id', flat=True)
    
    # 3. QuerySet Base
    recepciones_lista = DetalleRecepcion.objects.exclude(es_saldo_inicial=True).select_related('material', 'reporte').filter(id__in=ids_unicos)

    # 4. Aplicar filtros condicionales
    if f_base:
        recepciones_lista = recepciones_lista.filter(departamento__icontains=f_base)
    if f_em:
        recepciones_lista = recepciones_lista.filter(nro_control_entrada__icontains=f_em)
    if f_fecha_rep:
        recepciones_lista = recepciones_lista.filter(reporte__fecha_recepcion__icontains=f_fecha_rep)
    if f_fecha_ent:
        recepciones_lista = recepciones_lista.filter(fecha_recepcion__icontains=f_fecha_ent)
    if f_odc:
        recepciones_lista = recepciones_lista.filter(nro_odc__icontains=f_odc)
    if f_nota:
        recepciones_lista = recepciones_lista.filter(nro_nota_entrega__icontains=f_nota)
    if f_prov:
        recepciones_lista = recepciones_lista.filter(proveedor__icontains=f_prov)
    if f_mat:
        recepciones_lista = recepciones_lista.filter(
            Q(material__codigo__icontains=f_mat) | Q(material__descripcion__icontains=f_mat) | Q(descripcion_entrada__icontains=f_mat)
        )
    if f_obs:
        recepciones_lista = recepciones_lista.filter(observaciones__icontains=f_obs)
    if f_rm:
        recepciones_lista = recepciones_lista.filter(reporte__nro_reporte__icontains=f_rm)
    if f_vol:
        recepciones_lista = recepciones_lista.filter(volumen_carpeta__icontains=f_vol)

    recepciones_lista = recepciones_lista.order_by('-fecha_recepcion', '-id')

    # Anotación del total agrupado
    from django.db.models import OuterRef, Subquery, Sum, F, DecimalField
    
    totales_qs = DetalleRecepcion.objects.filter(
        nro_control_entrada=OuterRef('nro_control_entrada')
    ).values('nro_control_entrada').annotate(
        total=Sum(F('cantidad_recibida') * F('precio_unitario'), output_field=DecimalField())
    ).values('total')

    recepciones_lista = recepciones_lista.annotate(costo_total_agrupado=Subquery(totales_qs))

    # Filtrado por tipos para las pestañas
    entradas_em = recepciones_lista.filter(nro_control_entrada__startswith='EM')
    entradas_ea = recepciones_lista.filter(nro_control_entrada__startswith='EA')
    entradas_edg = recepciones_lista.filter(nro_control_entrada__startswith='EDG')
        
    # Paginación (Usaremos una paginación simple que afecte a la pestaña activa)
    # Para simplificar y cumplir el requerimiento de las 3 listas, las pasamos filtradas.
    from django.core.paginator import Paginator
    page_number = request.GET.get('page')
    
    recepciones_em = Paginator(entradas_em, 50).get_page(page_number)
    recepciones_ea = Paginator(entradas_ea, 50).get_page(page_number)
    recepciones_edg = Paginator(entradas_edg, 50).get_page(page_number)

    # 6. Preservar estado
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    query_prefix = query_params.urlencode() + '&' if query_params else ''

    contexto = {
        'recepciones_em': recepciones_em,
        'recepciones_ea': recepciones_ea,
        'recepciones_edg': recepciones_edg,
        'query_prefix': query_prefix,
        'filtros': {
            'base': f_base, 'em': f_em, 'fecha_rep': f_fecha_rep, 'fecha_ent': f_fecha_ent,
            'odc': f_odc, 'nota': f_nota, 'prov': f_prov, 'mat': f_mat,
            'obs': f_obs, 'rm': f_rm, 'vol': f_vol
        }
    }
    return render(request, 'inventario/lista_entradas.html', contexto)

# Vista 4: Crear Reporte (RP-00X) y carga múltiple de ítems por carrito
@login_required(login_url='login')
@roles_requeridos(['Coordinador de Almacén', 'Especialista de Activos']) 
def crear_recepcion(request):
    from django.db import transaction
    import json
    from decimal import Decimal

    if request.method == 'POST':
        form = ReporteRecepcionForm(request.POST)
        carrito_json = request.POST.get('carrito_datos', '[]')

        try:
            items_carrito = json.loads(carrito_json)
        except json.JSONDecodeError:
            items_carrito = []

        # Lógica de detección de ítems (Arrays directos POST para EDG)
        codigos_edg = request.POST.getlist('codigo_edg[]')
        
        if form.is_valid() and (items_carrito or codigos_edg):
            with transaction.atomic():
                reporte = form.save()

                # A. Procesar ítems desde el Carrito (JSON)
                for item in items_carrito:
                    tipo_ingreso_raw = item.get('tipo_entrada', 'MATERIAL')
                    mapeo_tipos = {
                        'MATERIAL': 'Material',
                        'ACTIVOS': 'Activo',
                        'DIRECTO AL GASTO': 'Directo al Gasto'
                    }
                    tipo_ingreso = mapeo_tipos.get(tipo_ingreso_raw, 'Material')
                    
                    material_obj = None
                    activo_obj = None
                    gasto_directo_obj = None
                    cant_recibida = Decimal(item.get('cantidad_recibida') or '0')

                    # --- LÓGICA DIRECTO AL GASTO (EDG) ---
                    es_edg = item.get('es_edg') is True or str(item.get('es_edg')).lower() == 'true'
                    requiere_rp = item.get('requiere_rp', 'si') 

                    if es_edg:
                        tipo_ingreso = 'Directo al Gasto'
                        if requiere_rp == 'si':
                            material_obj = None
                        else:
                            cod_edg = item.get('codigo_edg', '').strip().upper()
                            desc_edg = item.get('descripcion_edg', '').strip().upper()
                            um_edg = item.get('um_edg', '').strip().upper()
                            cargo_edg = item.get('cargo_edg', 'OPERACIONES')

                            if cod_edg and desc_edg:
                                gasto_directo_obj, _ = GastoDirecto.objects.get_or_create(
                                    codigo_dg=cod_edg,
                                    defaults={
                                        'descripcion': desc_edg,
                                        'unidad_medida_dg': um_edg,
                                        'cargo_dg': cargo_edg,
                                        'stock': 0 # Stock pasajero
                                    }
                                )
                    elif tipo_ingreso == 'Material':
                        material_id = item.get('material_id')
                        if material_id:
                            material_obj = Material.objects.get(id=material_id)
                        else:
                            codigo_material = item.get('material', '').split(' - ')[0].replace('[MATERIAL]', '').replace('[ACTIVOS]', '').replace('[DIRECTO AL GASTO]', '').strip()
                            material_obj = Material.objects.filter(codigo=codigo_material).first()
                    elif tipo_ingreso == 'Activo':
                        activo_id = item.get('activo_id')
                        if activo_id:
                            activo_obj = Activo.objects.get(id=activo_id)

                    import re
                    patron_odc = r'^PRSV-\d{4}-\d{10}$'
                    nro_odc_item = item.get('nro_odc', '').strip()
                    if not re.match(patron_odc, nro_odc_item) and not DetalleRecepcion.objects.filter(nro_odc=nro_odc_item).exists():
                        from django.core.exceptions import ValidationError
                        raise ValidationError(f"Error: La ODC '{nro_odc_item}' no cumple el formato nuevo ni existe en el histórico.")

                    detalle = DetalleRecepcion(
                        reporte=reporte,
                        material=material_obj,
                        activo=activo_obj,
                        gasto_directo=gasto_directo_obj,
                        tipo_ingreso=tipo_ingreso,
                        nro_odc=nro_odc_item,
                        fecha_recepcion=reporte.fecha_recepcion,
                        nro_rq=item.get('nro_rq'),
                        departamento=item.get('departamento'),
                        proveedor=item.get('proveedor'),
                        moneda=item.get('moneda', 'USD'),
                        eta=item.get('eta') or None,
                        nro_nota_entrega=item.get('nro_nota_entrega'),
                        cantidad_solicitada=Decimal(item.get('cantidad_solicitada') or '0'),
                        cantidad_recibida=cant_recibida,
                        precio_unitario=Decimal(item.get('precio_unitario') or '0'),
                        observaciones=item.get('observaciones')
                    )
                    detalle._tipo_entrada_manual = 'DIRECTO AL GASTO' if es_edg else item.get('tipo_entrada')
                    detalle.save()

                # B. Procesar ítems desde Arrays Directos (POST dinámico EDG)
                if codigos_edg:
                    descripciones_edg = request.POST.getlist('descripcion_edg[]')
                    ums_edg = request.POST.getlist('um_edg[]')
                    cargos_edg = request.POST.getlist('cargo_edg[]')
                    cantidades_edg = request.POST.getlist('cantidad_edg[]')
                    precios_edg = request.POST.getlist('precio_edg[]')
                    odcs_edg = request.POST.getlist('odc_edg[]')
                    notas_edg = request.POST.getlist('nota_edg[]')

                    for cod, desc, um, cargo, cant, precio, odc, nota in zip(
                        codigos_edg, descripciones_edg, ums_edg, cargos_edg, 
                        cantidades_edg, precios_edg, odcs_edg, notas_edg
                    ):
                        if not cod or not desc: continue
                        
                        gd_obj, _ = GastoDirecto.objects.get_or_create(
                            codigo_dg=cod.strip().upper(),
                            defaults={
                                'descripcion': desc.strip().upper(),
                                'unidad_medida_dg': um.strip().upper(),
                                'cargo_dg': cargo.strip().upper(),
                                'stock': 0
                            }
                        )

                        detalle = DetalleRecepcion(
                            reporte=reporte,
                            gasto_directo=gd_obj,
                            tipo_ingreso='Directo al Gasto',
                            nro_odc=odc.strip().upper(),
                            fecha_recepcion=reporte.fecha_recepcion,
                            cantidad_recibida=Decimal(cant or '0'),
                            precio_unitario=Decimal(precio or '0'),
                            nro_nota_entrega=nota.strip().upper()
                        )
                        detalle._tipo_entrada_manual = 'DIRECTO AL GASTO'
                        detalle.save()

                registrar_actividad(request.user, "Registró Recepción", f"Reporte {reporte.nro_reporte} del día {reporte.fecha_recepcion.strftime('%d/%m/%Y')}")
            return redirect('lista_entradas')
        else:
            form = ReporteRecepcionForm(request.POST)
    else:
        form = ReporteRecepcionForm()

    form_detalle = DetalleRecepcionForm()
    odcs_existentes = list(DetalleRecepcion.objects.exclude(nro_odc__isnull=True).exclude(nro_odc__exact='').values_list('nro_odc', flat=True).distinct())

    contexto = {
        'form': form,
        'form_detalle': form_detalle,
        'odcs_existentes': odcs_existentes,
        'materiales': Material.objects.all(),
        'activos': Activo.objects.all(),
    }
    return render(request, 'inventario/crear_recepcion.html', contexto)

# Vista 4B: Formulario independiente para registrar entradas (EM)
@login_required(login_url='login')
@roles_requeridos(['Coordinador de Almacén', 'Especialista de Activos', 'Analista de Almacén']) 
def registrar_entrada(request):
    from django.db import transaction
    import json
    from decimal import Decimal

    if request.method == 'POST':
        # 1. Validar Firma Digital (Contraseña)
        password_firma = request.POST.get('password_firma', '').strip()
        if not password_firma or not request.user.check_password(password_firma):
            from django.contrib import messages
            messages.error(request, "Firma digital inválida: La contraseña ingresada es incorrecta. Transacción cancelada.")
            return redirect('registrar_entrada')

        carrito_json = request.POST.get('carrito_datos', '[]')
        
        try:
            items_carrito = json.loads(carrito_json)
        except json.JSONDecodeError:
            items_carrito = []

        # --- LÓGICA DE CAPTURA PARA EDG (Arrays directos) ---
        codigos_edg = request.POST.getlist('codigo_edg[]')

        if items_carrito or codigos_edg:
            with transaction.atomic():
                reporte_id = request.POST.get('reporte_id')
                reporte_obj = None
                if reporte_id:
                    reporte_obj = ReporteRecepcion.objects.filter(id=reporte_id).first()

                # A. Procesar ítems desde el Carrito (JSON)
                for item in items_carrito:
                    tipo_ingreso_raw = item.get('tipo_entrada', 'MATERIAL')
                    mapeo_tipos = {
                        'MATERIAL': 'Material',
                        'ACTIVOS': 'Activo',
                        'DIRECTO AL GASTO': 'Directo al Gasto'
                    }
                    tipo_ingreso = mapeo_tipos.get(tipo_ingreso_raw, 'Material')
                    
                    # Captura universal de cantidades y precios del ítem
                    cant_solicitada_item = Decimal(item.get('cantidad_solicitada') or '0')
                    cant_recibida_item = Decimal(item.get('cantidad_recibida') or '0')
                    precio_unitario_item = Decimal(item.get('precio_unitario') or '0')

                    material_obj = None
                    activo_obj = None
                    gasto_directo_obj = None

                    if tipo_ingreso == 'Directo al Gasto':
                        cod_edg = item.get('codigo_edg', '').strip().upper()
                        desc_edg = item.get('descripcion_edg', '').strip().upper()
                        um_edg = item.get('um_edg', '').strip().upper()
                        cargo_edg = item.get('cargo_edg', 'OPERACIONES')

                        if cod_edg and desc_edg:
                            gasto_directo_obj, _ = GastoDirecto.objects.get_or_create(
                                codigo_dg=cod_edg,
                                defaults={
                                    'descripcion': desc_edg,
                                    'unidad_medida_dg': um_edg,
                                    'cargo_dg': cargo_edg,
                                    'stock': 0 # Stock pasajero
                                }
                            )
                    elif tipo_ingreso == 'Material':
                        material_id = item.get('material_id')
                        if material_id:
                            material_obj = Material.objects.get(id=material_id)
                        else:
                            codigo_material = item.get('material', '').split(' - ')[0].replace('[MATERIAL]', '').replace('[ACTIVOS]', '').replace('[DIRECTO AL GASTO]', '').strip()
                            material_obj = Material.objects.filter(codigo=codigo_material).first()
                    elif tipo_ingreso == 'Activo':
                        activo_id = item.get('activo_id')
                        if activo_id:
                            activo_obj = Activo.objects.get(id=activo_id)

                    rep_final = reporte_obj
                    if rep_final is None:
                        nro_odc_item = item.get('nro_odc', '').strip()
                        nro_nota_item = item.get('nro_nota_entrega', '').strip()
                        
                        if nro_odc_item and nro_nota_item:
                            sibling = DetalleRecepcion.objects.filter(
                                nro_odc=nro_odc_item,
                                nro_nota_entrega=nro_nota_item,
                                reporte__isnull=False
                            ).select_related('reporte').first()
                            if sibling:
                                rep_final = sibling.reporte
                        elif nro_odc_item:
                            sibling = DetalleRecepcion.objects.filter(
                                nro_odc=nro_odc_item,
                                reporte__isnull=False
                            ).select_related('reporte').first()
                            if sibling:
                                rep_final = sibling.reporte

                    fecha_entrada_str = item.get('fecha_entrada')
                    import datetime as dt
                    if fecha_entrada_str:
                        try:
                            fecha_para_codigo = dt.date.fromisoformat(fecha_entrada_str)
                        except ValueError:
                            fecha_para_codigo = timezone.now().date()
                    else:
                        fecha_para_codigo = timezone.now().date()

                    import re
                    patron_odc = r'^PRSV-\d{4}-\d{10}$'
                    nro_odc_item = item.get('nro_odc', '').strip()
                    if not re.match(patron_odc, nro_odc_item) and not DetalleRecepcion.objects.filter(nro_odc=nro_odc_item).exists():
                        from django.core.exceptions import ValidationError
                        raise ValidationError(f"Error: La ODC '{nro_odc_item}' no cumple el formato nuevo ni existe en el historico.")

                    detalle = DetalleRecepcion(
                        reporte=rep_final,
                        material=material_obj,
                        activo=activo_obj,
                        gasto_directo=gasto_directo_obj,
                        tipo_ingreso=tipo_ingreso,
                        descripcion_entrada=item.get('descripcion_entrada') or item.get('material_texto'),
                        nro_odc=nro_odc_item,
                        fecha_recepcion=fecha_para_codigo,
                        nro_rq=item.get('nro_rq'),
                        departamento=item.get('base'),
                        proveedor=item.get('proveedor'),
                        moneda=item.get('moneda', 'USD'),
                        eta=item.get('eta') or None,
                        nro_nota_entrega=item.get('nro_nota_entrega'),
                        cantidad_solicitada=cant_solicitada_item,
                        cantidad_recibida=cant_recibida_item,
                        precio_unitario=precio_unitario_item,
                        observaciones=item.get('observaciones')
                    )

                    detalle._tipo_entrada_manual = item.get('tipo_entrada')
                    detalle.save()

                # B. Procesar ítems desde Arrays Directos (EDG)
                if codigos_edg:
                    descripciones_edg = request.POST.getlist('descripcion_edg[]')
                    ums_edg = request.POST.getlist('um_edg[]')
                    cargos_edg = request.POST.getlist('cargo_edg[]')
                    cantidades_edg = request.POST.getlist('cantidad_edg[]')
                    precios_edg = request.POST.getlist('precio_edg[]')
                    odcs_edg = request.POST.getlist('odc_edg[]')
                    notas_edg = request.POST.getlist('nota_edg[]')

                    for cod, desc, um, cargo, cant, precio, odc, nota in zip(
                        codigos_edg, descripciones_edg, ums_edg, cargos_edg, 
                        cantidades_edg, precios_edg, odcs_edg, notas_edg
                    ):
                        if not cod or not desc: continue
                        
                        gd_obj, _ = GastoDirecto.objects.get_or_create(
                            codigo_dg=cod.strip().upper(),
                            defaults={
                                'descripcion': desc.strip().upper(),
                                'unidad_medida_dg': um.strip().upper(),
                                'cargo_dg': cargo.strip().upper(),
                                'stock': 0
                            }
                        )

                        # Intentar buscar reporte asociado a la ODC
                        rep_item = reporte_obj
                        if not rep_item and odc:
                            sibling = DetalleRecepcion.objects.filter(nro_odc=odc.strip().upper(), reporte__isnull=False).first()
                            if sibling:
                                rep_item = sibling.reporte

                        detalle = DetalleRecepcion(
                            reporte=rep_item,
                            gasto_directo=gd_obj,
                            tipo_ingreso='Directo al Gasto',
                            nro_odc=odc.strip().upper(),
                            fecha_recepcion=timezone.now().date(),
                            cantidad_recibida=Decimal(cant or '0'),
                            precio_unitario=Decimal(precio or '0'),
                            nro_nota_entrega=nota.strip().upper()
                        )
                        detalle._tipo_entrada_manual = 'DIRECTO AL GASTO'
                        detalle.save()

                # Auditoría: Registrar movimiento
                cant_total = len(items_carrito) + len(codigos_edg)
                registrar_actividad(request.user, "Registró Entrada", f"Ingreso de {cant_total} ítems de material a almacén.")
            return redirect('lista_entradas')
    form_detalle = DetalleRecepcionForm()
    odcs_existentes = list(DetalleRecepcion.objects.exclude(nro_odc__isnull=True).exclude(nro_odc__exact='').values_list('nro_odc', flat=True).distinct())
    reportes_recientes = ReporteRecepcion.objects.all().order_by('-fecha_recepcion', '-id')[:30]
    departamentos = PresupuestoAnual.objects.values_list('departamento', flat=True).distinct().order_by('departamento')

    contexto = {
        'form_detalle': form_detalle,
        'odcs_existentes': odcs_existentes,
        'reportes_recientes': reportes_recientes,
        'materiales': Material.objects.all(),
        'activos': Activo.objects.all(),
        'departamentos': departamentos,
    }
    return render(request, 'inventario/registrar_entrada.html', contexto)

# Vista 5: Llenar el Reporte con Ítems (EM26001)
@login_required(login_url='login')
def detalle_recepcion(request, reporte_id):
    reporte = get_object_or_404(ReporteRecepcion, id=reporte_id)
    
    # --- LÓGICA DEL FILTRO ---
    # Si la URL dice ?filtro=diferencias, filtramos la lista
    filtro_activo = request.GET.get('filtro')
    
    if filtro_activo == 'diferencias':
        from django.db.models import F
        # Trae solo los ítems donde lo recibido NO es igual a lo solicitado
        items = DetalleRecepcion.objects.filter(reporte=reporte).exclude(cantidad_solicitada=F('cantidad_recibida')).order_by('-id')
    else:
        # Trae todos normalmente
        items = DetalleRecepcion.objects.filter(reporte=reporte).order_by('-id')

    if request.method == 'POST':
        form = DetalleRecepcionForm(request.POST)
        if form.is_valid():
            nuevo_item = form.save(commit=False)
            nuevo_item.reporte = reporte
            nuevo_item.save()
            return redirect('lista_entradas')
    else:
        form = DetalleRecepcionForm()

    contexto = {
        'reporte': reporte,
        'items': items,
        'form': form,
        'filtro_activo': filtro_activo # Pasamos esto al HTML para saber si el botón está encendido
    }
    return render(request, 'inventario/detalle_recepcion.html', contexto)

@login_required(login_url='login')
@user_passes_test(es_almacenista, login_url='lista_entradas')
def editar_entrada(request, pk):
    entrada = get_object_or_404(DetalleRecepcion, pk=pk)
    
    if request.method == 'POST':
        # 1. Validar Firma Digital (Contraseña)
        password_firma = request.POST.get('password_firma', '').strip()
        if not password_firma or not request.user.check_password(password_firma):
            from django.contrib import messages
            messages.error(request, "Firma digital inválida: La contraseña ingresada es incorrecta. Edición cancelada.")
            return redirect('lista_entradas')
            
        form = DetalleRecepcionEditForm(request.POST, instance=entrada)
        if form.is_valid():
            form.save()
            from django.contrib import messages
            messages.success(request, f"Entrada {entrada.nro_control_entrada} actualizada correctamente.")
            return redirect('lista_entradas')
    else:
        form = DetalleRecepcionEditForm(instance=entrada)
    
    contexto = {
        'form': form,
        'entrada': entrada
    }
    return render(request, 'inventario/editar_entrada.html', contexto)

@login_required(login_url='login')
@user_passes_test(es_almacenista, login_url='lista_entradas')
def eliminar_entrada(request, pk):
    entrada = get_object_or_404(DetalleRecepcion, pk=pk)
    
    if request.method == 'POST':
        # 1. Validar Firma Digital (Contraseña)
        password_firma = request.POST.get('password_firma', '').strip()
        if not password_firma or not request.user.check_password(password_firma):
            from django.contrib import messages
            messages.error(request, "Firma digital inválida: La contraseña ingresada es incorrecta. Eliminación cancelada.")
            return redirect('lista_entradas')
            
        try:
            with transaction.atomic():
                # Restauración inversa de stock (Validación condicional Material/Activo)
                if entrada.material:
                    entrada.material.stock_actual -= entrada.cantidad_recibida
                    entrada.material.save()
                elif entrada.activo:
                    # Los activos manejan stock como Integer
                    entrada.activo.stock -= int(entrada.cantidad_recibida)
                    entrada.activo.save()
                
                nro_em = entrada.nro_control_entrada
                entrada.delete()
                
            from django.contrib import messages
            messages.success(request, f"Entrada {nro_em} eliminada correctamente. Stock descontado.")
        except Exception as e:
            print(f"ERROR CRÍTICO AL ELIMINAR ENTRADA: {e}")
            from django.contrib import messages
            messages.error(request, f"Error al eliminar: {str(e)}")
            
    return redirect('lista_entradas')

# VISTA 6: Lista de Despachos (RIM)
@login_required(login_url='login')
def lista_salidas(request):
    # 1. Capturar filtros (TODOS)
    f_fecha = request.GET.get('f_fecha', '').strip()
    f_rim = request.GET.get('f_rim', '').strip()
    f_mat = request.GET.get('f_mat', '').strip()
    f_cant = request.GET.get('f_cant', '').strip()
    f_um = request.GET.get('f_um', '').strip()
    f_depto = request.GET.get('f_depto', '').strip()
    f_cc = request.GET.get('f_cc', '').strip()

    # 2. QuerySet Base
    salidas_qs = SalidaMaterial.objects.exclude(
        Q(nro_rim__startswith='AJUSTE-MIG') | Q(departamento='MIGRACIÓN')
    ).select_related('material', 'activo', 'centro_costo').all()

    # 3. Filtros
    if f_fecha:
        salidas_qs = salidas_qs.filter(fecha_despacho__icontains=f_fecha)
    if f_rim:
        salidas_qs = salidas_qs.filter(nro_rim__icontains=f_rim)
    if f_mat:
        salidas_qs = salidas_qs.filter(
            Q(material__codigo__icontains=f_mat) | Q(material__descripcion__icontains=f_mat)
        )
    if f_cant:
        salidas_qs = salidas_qs.filter(cantidad__icontains=f_cant)
    if f_um:
        salidas_qs = salidas_qs.filter(material__unidad_medida__icontains=f_um)
    if f_depto:
        salidas_qs = salidas_qs.filter(departamento__icontains=f_depto)
    if f_cc:
        salidas_qs = salidas_qs.filter(centro_costo__icontains=f_cc)

    salidas_qs = salidas_qs.order_by('-fecha_despacho', '-id')

    paginator = Paginator(salidas_qs, 50)
    page_number = request.GET.get('page')
    salidas_paginadas = paginator.get_page(page_number)

    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    query_prefix = query_params.urlencode() + '&' if query_params else ''

    contexto = {
        'salidas': salidas_paginadas,
        'query_prefix': query_prefix,
        'filtros': {
            'fecha': f_fecha, 'rim': f_rim, 'mat': f_mat, 'cant': f_cant,
            'um': f_um, 'depto': f_depto, 'cc': f_cc
        }
    }
    return render(request, 'inventario/lista_salidas.html', contexto)

# VISTA 7: Registrar un Nuevo Despacho (RIM)
@login_required(login_url='login')
@user_passes_test(es_almacenista, login_url='lista_entradas') # <--- ESCUDO NUEVO
def crear_salida(request):
    if request.method == 'POST':
        import json
        from decimal import Decimal
        from django.http import JsonResponse
        
        try:
            print(f"DEBUG: Recibida petición POST en crear_salida. User: {request.user}")
            data = json.loads(request.body)
            items_carrito = data.get('items', [])
            necesita_guia = data.get('necesita_guia', False)
            print(f"DEBUG: Items en carrito: {len(items_carrito)}. Necesita Guía: {necesita_guia}")
        except Exception as e:
            print(f"DEBUG ERROR: Fallo al parsear JSON: {str(e)}")
            return JsonResponse({'status': 'error', 'message': f'Datos JSON inválidos: {str(e)}'}, status=400)

        if items_carrito:
            try:
                with transaction.atomic():
                    for index, item in enumerate(items_carrito):
                        tipo_s = item.get('tipo_salida', 'SM')
                        material_id = item.get('material_id')
                        
                        material_obj = None
                        activo_obj = None

                        if tipo_s == 'SA':
                            # Buscar en Maestro de Activos
                            try:
                                activo_obj = Activo.objects.get(id=material_id)
                            except Activo.DoesNotExist:
                                return JsonResponse({'status': 'error', 'message': f"El activo con ID {material_id} no existe en el maestro."}, status=400)
                        else:
                            # Buscar en Maestro de Materiales
                            try:
                                material_obj = Material.objects.get(id=material_id)
                            except Material.DoesNotExist:
                                return JsonResponse({'status': 'error', 'message': f"El material con ID {material_id} no existe en el maestro."}, status=400)
                        
                        # Creamos la salida individual
                        nueva_salida = SalidaMaterial(
                            material=material_obj,
                            activo=activo_obj,
                            fecha_despacho=item['fecha_despacho'],
                            nro_rim=item['nro_rim'],
                            cantidad=Decimal(item['cantidad']),
                            tipo_salida=tipo_s,
                            numero_salida_correlativo=item.get('numero_salida_correlativo'),
                            departamento=item.get('departamento'),
                            centro_costo_id=item.get('centro_costo') if item.get('centro_costo') and str(item.get('centro_costo')).isdigit() else None,
                            cuenta_contable=item.get('cuenta_contable'),
                            descripcion_cuenta=item.get('descripcion_cuenta'),
                            partida_presupuestaria=item.get('partida_presupuestaria'),
                            rubro_1=item.get('rubro_1'),
                            rubro_2=item.get('rubro_2'),
                            creado_por=request.user
                        )
                        
                        print(f"DEBUG: Intentando salvar SalidaMaterial para RIM {nueva_salida.nro_rim}...")
                        nueva_salida.save()
                        print(f"DEBUG: Éxito al guardar ítem {index + 1}. Stock descontado.")

                # Registrar actividad de auditoría
                rim_ejemplo = items_carrito[0].get('nro_rim') if items_carrito else 'N/A'
                registrar_actividad(request.user, "Registró Salida", f"Salida de {len(items_carrito)} ítems. RIM: {rim_ejemplo}")

                # Respuesta de éxito con URL de redirección
                from django.urls import reverse
                redirect_url = reverse('crear_guia') if necesita_guia else reverse('lista_salidas')
                print(f"DEBUG: Proceso completo exitoso. Redirigiendo a: {redirect_url}")
                return JsonResponse({
                    'status': 'ok', 
                    'message': 'Despacho procesado con éxito',
                    'redirect_url': redirect_url
                })

            except Exception as e:
                import traceback
                print(f"DEBUG CRITICAL ERROR: {str(e)}")
                traceback.print_exc()
                return JsonResponse({'status': 'error', 'message': f"Error de base de datos: {str(e)}"}, status=400)
        else:
            print("DEBUG: Carrito vacío recibido.")
            return JsonResponse({'status': 'error', 'message': 'El carrito está vacío'}, status=400)



    else:
        form = SalidaMaterialForm()
    
    materiales = Material.objects.all().order_by('codigo')
    activos = Activo.objects.all().order_by('codigo_activo')

    contexto = {
        'form': form,
        'materiales': materiales,
        'activos': activos
    }
    return render(request, 'inventario/crear_salida.html', contexto)

# VISTA 8: Generar PDF de la Nota de Despacho (RIM)
@login_required(login_url='login')
def generar_pdf_salida(request, salida_id):
    from django.template.loader import render_to_string
    from weasyprint import HTML
    
    # 1. Buscamos el despacho específico
    salida_base = get_object_or_404(SalidaMaterial, id=salida_id)
    
    # 2. Si el usuario quiere ver TODO el RIM agrupado (todos los materiales del mismo nro_rim)
    salidas_agrupadas = SalidaMaterial.objects.filter(
        nro_rim=salida_base.nro_rim,
        fecha_despacho=salida_base.fecha_despacho
    ).prefetch_related('detalles__detalle_recepcion')
    
    # 3. Contexto
    template_path = 'inventario/pdf_salida.html'
    context = {
        'salida': salida_base,
        'salidas_agrupadas': salidas_agrupadas,
    }
    
    # 4. Renderizamos
    html_string = render_to_string(template_path, context, request=request)
    
    # 5. Generamos PDF con WeasyPrint
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    pdf = html.write_pdf()
    
    # 6. Respuesta
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="RIM_{salida_base.nro_rim}.pdf"'
    
    return response

@login_required(login_url='login')
@user_passes_test(es_almacenista, login_url='lista_salidas')
def editar_salida(request, pk):
    salida = get_object_or_404(SalidaMaterial, pk=pk)
    
    if request.method == 'POST':
        form = SalidaMaterialEditForm(request.POST, instance=salida)
        if form.is_valid():
            form.save()
            from django.contrib import messages
            messages.success(request, f"Registro de Despacho {salida.nro_rim} actualizado correctamente.")
            return redirect('lista_salidas')
    else:
        form = SalidaMaterialEditForm(instance=salida)
    
    contexto = {
        'form': form,
        'salida': salida
    }
    return render(request, 'inventario/editar_salida.html', contexto)

@login_required(login_url='login')
@user_passes_test(es_almacenista, login_url='lista_salidas')
def eliminar_salida(request, pk):
    salida = get_object_or_404(SalidaMaterial, pk=pk)
    
    try:
        with transaction.atomic():
            # 1. Restaurar el stock al catálogo correspondiente (Reversión de Salida)
            if salida.material:
                salida.material.stock_actual += salida.cantidad
                salida.material.save()
            elif salida.activo:
                # Los activos manejan stock como Integer
                salida.activo.stock += int(salida.cantidad)
                salida.activo.save()
            
            # 2. Eliminar el registro de salida
            nro_rim = salida.nro_rim
            salida.delete()
            
        from django.contrib import messages
        messages.success(request, f"Despacho {nro_rim} eliminado. Stock restaurado al maestro.")
    except Exception as e:
        print(f"ERROR CRÍTICO AL ELIMINAR SALIDA: {e}")
        from django.contrib import messages
        messages.error(request, f"Error al eliminar despacho: {str(e)}")
        
    return redirect('lista_salidas')

# VISTA 9: Lista de Guías de Traslado y Transferencia
@login_required(login_url='login')
def lista_guias(request):
    # 1. Parámetros de Filtro por columna
    f_nro_guia = request.GET.get('f_nro_guia', '').strip()
    f_fecha = request.GET.get('f_fecha', '').strip()
    f_destino = request.GET.get('f_destino', '').strip()
    f_conductor = request.GET.get('f_conductor', '').strip()
    active_tab = request.GET.get('tab', 'traslado')

    # 2. Querysets Base
    # Guías que contienen al menos un Material (Traslados)
    guias_traslado_qs = GuiaTraslado.objects.filter(
        salidas__material__isnull=False
    ).distinct()
    
    # Guías que contienen al menos un Activo Fijo (Transferencias)
    guias_transferencia_qs = GuiaTraslado.objects.filter(
        salidas__activo__isnull=False
    ).distinct()

    # 3. Aplicar Filtros Dinámicos
    if f_nro_guia:
        guias_traslado_qs = guias_traslado_qs.filter(nro_guia__icontains=f_nro_guia)
        guias_transferencia_qs = guias_transferencia_qs.filter(nro_guia__icontains=f_nro_guia)
    
    if f_fecha:
        guias_traslado_qs = guias_traslado_qs.filter(fecha__icontains=f_fecha)
        guias_transferencia_qs = guias_transferencia_qs.filter(fecha__icontains=f_fecha)
        
    if f_destino:
        guias_traslado_qs = guias_traslado_qs.filter(taladro_destino__icontains=f_destino)
        guias_transferencia_qs = guias_transferencia_qs.filter(taladro_destino__icontains=f_destino)
        
    if f_conductor:
        guias_traslado_qs = guias_traslado_qs.filter(
            Q(conductor__icontains=f_conductor) |
            Q(placa__icontains=f_conductor) |
            Q(ci_conductor__icontains=f_conductor)
        )
        guias_transferencia_qs = guias_transferencia_qs.filter(
            Q(conductor__icontains=f_conductor) |
            Q(placa__icontains=f_conductor) |
            Q(ci_conductor__icontains=f_conductor)
        )

    # Ordenamiento final
    guias_traslado_qs = guias_traslado_qs.order_by('-fecha', '-id')
    guias_transferencia_qs = guias_transferencia_qs.order_by('-fecha', '-id')

    # 4. Paginación Server-side (10 registros por página)
    paginator_traslado = Paginator(guias_traslado_qs, 10)
    page_traslado = request.GET.get('page_traslado')
    page_obj_traslado = paginator_traslado.get_page(page_traslado)

    paginator_transferencia = Paginator(guias_transferencia_qs, 10)
    page_transferencia = request.GET.get('page_transferencia')
    page_obj_transferencia = paginator_transferencia.get_page(page_transferencia)

    # Conservar todos los parámetros de consulta para los links de paginación
    query_params_traslado = request.GET.copy()
    if 'page_traslado' in query_params_traslado:
        del query_params_traslado['page_traslado']
    query_prefix_traslado = query_params_traslado.urlencode() + '&' if query_params_traslado else ''

    query_params_transferencia = request.GET.copy()
    if 'page_transferencia' in query_params_transferencia:
        del query_params_transferencia['page_transferencia']
    query_prefix_transferencia = query_params_transferencia.urlencode() + '&' if query_params_transferencia else ''

    contexto = {
        'guias_traslado': page_obj_traslado,
        'guias_transferencia': page_obj_transferencia,
        'query_prefix_traslado': query_prefix_traslado,
        'query_prefix_transferencia': query_prefix_transferencia,
        'active_tab': active_tab,
        'filtros': {
            'nro_guia': f_nro_guia,
            'fecha': f_fecha,
            'destino': f_destino,
            'conductor': f_conductor,
        }
    }
    return render(request, 'inventario/lista_guias.html', contexto)

# VISTA 10: Crear el encabezado de la Guía (El camión)
@login_required(login_url='login')
@user_passes_test(es_almacenista, login_url='lista_entradas') # <--- ESCUDO NUEVO
def crear_guia(request):
    if request.method == 'POST':
        form = GuiaTrasladoForm(request.POST)
        if form.is_valid():
            guia = form.save()
            # Al guardar, lo enviamos directo a la pantalla para meterle los materiales
            return redirect('detalle_guia', guia_id=guia.id) 
    else:
        form = GuiaTrasladoForm()
    return render(request, 'inventario/crear_guia.html', {'form': form})

@login_required(login_url='login')
@user_passes_test(es_almacenista, login_url='lista_guias')
def editar_guia(request, pk):
    guia = get_object_or_404(GuiaTraslado, pk=pk)
    if request.method == 'POST':
        form = GuiaTrasladoForm(request.POST, instance=guia)
        if form.is_valid():
            form.save()
            return redirect('lista_guias')
    else:
        form = GuiaTrasladoForm(instance=guia)
    
    return render(request, 'inventario/crear_guia.html', {
        'form': form,
        'edit_mode': True,
        'guia': guia
    })

@login_required(login_url='login')
@user_passes_test(es_almacenista, login_url='lista_guias')
def eliminar_guia(request, pk):
    guia = get_object_or_404(GuiaTraslado, pk=pk)
    if request.method == 'POST':
        guia.delete()
        return redirect('lista_guias')
    return redirect('lista_guias')


# VISTA 11: Armar la Guía (La magia de los checkboxes)
@login_required(login_url='login')
def detalle_guia(request, guia_id):
    guia = get_object_or_404(GuiaTraslado, id=guia_id)
    
    # 1. Traemos los materiales que YA ESTÁN en este camión
    items_en_guia = SalidaMaterial.objects.filter(guia=guia)
    
    # 2. Traemos los RIMs que están "Huérfanos" (Excluyendo ajustes de migración fantasma)
    items_pendientes = SalidaMaterial.objects.filter(
        guia__isnull=True
    ).exclude(
        nro_rim__startswith='AJUSTE-MIG-'
    ).order_by('-fecha_despacho')

    if request.method == 'POST':
        # Recibimos la lista de los IDs que el usuario marcó con el Checkbox (✔)
        ids_seleccionados = request.POST.getlist('rims_seleccionados')
        if ids_seleccionados:
            # Actualizamos esos RIMs en la base de datos para decirles: "Ahora pertenecen a esta Guía"
            SalidaMaterial.objects.filter(id__in=ids_seleccionados).update(guia=guia)
        
        return redirect('detalle_guia', guia_id=guia.id)

    contexto = {
        'guia': guia,
        'items_en_guia': items_en_guia,
        'items_pendientes': items_pendientes
    }
    return render(request, 'inventario/detalle_guia.html', contexto)


# ==================================================
# FLUJO: GUÍAS DE TRANSFERENCIA (SOLO ACTIVOS)
# ==================================================

@login_required(login_url='login')
@user_passes_test(es_almacenista, login_url='lista_guias')
def crear_guia_transferencia(request):
    if request.method == 'POST':
        form = GuiaTrasladoForm(request.POST)
        if form.is_valid():
            guia = form.save()
            # Al guardar, lo enviamos directo a la pantalla para meterle los ACTIVOS
            return redirect('detalle_guia_transferencia', guia_id=guia.id) 
    else:
        form = GuiaTrasladoForm()
    return render(request, 'inventario/crear_guia_transferencia.html', {'form': form})


@login_required(login_url='login')
def detalle_guia_transferencia(request, guia_id):
    guia = get_object_or_404(GuiaTraslado, id=guia_id)
    
    # 1. Traemos los ACTIVOS que YA ESTÁN en este camión
    items_en_guia = SalidaMaterial.objects.filter(guia=guia, activo__isnull=False).select_related('activo')
    
    # 2. Traemos los ACTIVOS que están "Huérfanos"
    items_pendientes = SalidaMaterial.objects.filter(
        guia__isnull=True,
        activo__isnull=False
    ).exclude(
        nro_rim__startswith='AJUSTE-MIG-'
    ).select_related('activo').order_by('-fecha_despacho')

    if request.method == 'POST':
        # Recibimos la lista de los IDs que el usuario marcó con el Checkbox (✔)
        ids_seleccionados = request.POST.getlist('rims_seleccionados')
        if ids_seleccionados:
            # Actualizamos esos RIMs en la base de datos para decirles: "Ahora pertenecen a esta Guía"
            SalidaMaterial.objects.filter(id__in=ids_seleccionados).update(guia=guia)
        
        return redirect('detalle_guia_transferencia', guia_id=guia.id)

    contexto = {
        'guia': guia,
        'items_en_guia': items_en_guia,
        'items_pendientes': items_pendientes,
        'es_transferencia': True
    }
    return render(request, 'inventario/detalle_guia_transferencia.html', contexto)

@login_required(login_url='login')
def quitar_de_guia(request, item_id):
    """
    Desvincula un material (RIM) de una guía de traslado.
    """
    item = get_object_or_404(SalidaMaterial, id=item_id)
    guia_id = item.guia.id if item.guia else None
    
    # Desvinculamos
    item.guia = None
    item.save()
    
    # Soporte para AJAX
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('ajax') == '1':
        return JsonResponse({'status': 'ok', 'message': 'Material quitado de la guía'})
        
    if guia_id:
        return redirect('detalle_guia', guia_id=guia_id)
    return redirect('lista_guias')


# ==================================================
# VISTA: Generar PDF de la Guía de Traslado con WeasyPrint
# ==================================================

@login_required(login_url='login')
def generar_guia_pdf(request, pk):
    from django.template.loader import render_to_string
    from weasyprint import HTML
    # 1. Buscamos la guía específica
    guia = get_object_or_404(GuiaTraslado, pk=pk)
    
    # 2. Buscamos todos los ítems (RIMs) asociados a esta guía
    items = SalidaMaterial.objects.filter(guia=guia).order_by('id')
    items_list = list(items)
    
    # 3. Lógica de Chunking: Dividir en grupos de 12 para evitar que el diseño se rompa
    max_items_por_pagina = 12
    paginas = []
    
    for i in range(0, len(items_list), max_items_por_pagina):
        chunk = items_list[i : i + max_items_por_pagina]
        padding_count = max_items_por_pagina - len(chunk)
        paginas.append({
            'items': chunk,
            'padding': range(padding_count),
            'inicio_conteo': i + 1  # Para que el correlativo de ítems sea continuo
        })
    
    total_paginas = len(paginas)
    
    # 4. Le decimos qué plantilla de diseño usar
    template_path = 'inventario/guia_traslado_pdf.html'
    context = {
        'guia': guia, 
        'paginas': paginas,
        'total_paginas': total_paginas,
    }
    
    # 4. Renderizamos el HTML a string, pasando el request para resolver rutas
    html_string = render_to_string(template_path, context, request=request)
    
    # 5. Generamos el PDF con WeasyPrint
    # base_url permite que WeasyPrint encuentre los archivos estáticos usando la URL absoluta del servidor
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    pdf = html.write_pdf()
    
    # 6. Configuramos la respuesta HTTP
    response = HttpResponse(pdf, content_type='application/pdf')
    
    # inline indica al navegador que abra el PDF en una pestaña en lugar de descargarlo directamente
    response['Content-Disposition'] = f'inline; filename="Guia_{guia.nro_guia}.pdf"'
    
    return response


# ==================================================
# VISTA: Generar PDF de la Guía de Transferencia (Activos - ALM-FORM-006)
# ==================================================
@login_required(login_url='login')
def generar_pdf_transferencia(request, guia_id):
    from django.template.loader import render_to_string
    from weasyprint import HTML
    
    # 1. Obtener la guía (Cabecera)
    guia_obj = get_object_or_404(GuiaTraslado, id=guia_id)
    
    # 2. Obtener los activos asociados (Detalle de la carga)
    items_qs = SalidaMaterial.objects.filter(guia=guia_obj, activo__isnull=False).select_related('activo')
    
    # 3. Mapear datos de la guía para la plantilla (Compatibilidad ALM-FORM-006)
    guia_data = {
        'numero': guia_obj.nro_guia,
        'fecha': guia_obj.fecha,
        'hora': guia_obj.hora,
        'destino_nombre': guia_obj.taladro_destino,
        'destino_direccion_1': guia_obj.direccion,
        'destino_direccion_2': "", 
        'destino_ciudad': guia_obj.ciudad,
        'conductor_nombre': guia_obj.conductor,
        'conductor_ci': guia_obj.ci_conductor,
        'vehiculo_tipo': guia_obj.vehiculo,
        'vehiculo_color': guia_obj.color,
        'vehiculo_placa': guia_obj.placa,
        'vehiculo_marca': guia_obj.marca,
        'vehiculo_modelo': guia_obj.modelo,
        'entregado_nombre': guia_obj.nombre_entregado,
        'aprobado_nombre': guia_obj.nombre_aprobador or "PENDIENTE",
        'dsi_nombre': "", 
        'recibido_nombre': "",
    }
    
    # Dividir observaciones para las líneas del diseño
    obs = guia_obj.observaciones or ""
    obs_lines = obs.split('\n')
    guia_data['observacion_linea1'] = obs_lines[0] if len(obs_lines) > 0 else ""
    guia_data['observacion_linea2'] = obs_lines[1] if len(obs_lines) > 1 else ""
    guia_data['observacion_linea3'] = obs_lines[2] if len(obs_lines) > 2 else ""
    guia_data['observacion_linea4'] = obs_lines[3] if len(obs_lines) > 3 else ""

    # 4. Mapear datos de los ACTIVOS (Para las dos filas del HTML)
    activos_list = []
    for item in items_qs:
        activos_list.append({
            'codigo': item.activo.codigo_activo,
            'descripcion': item.activo.descripcion,
            'cantidad': int(item.cantidad),
            'um': item.activo.unidad_medida_ac,
            'marca': item.activo.marca or "-",
            'modelo': item.activo.modelo or "-",
            'serial': item.activo.serial or "-",
        })
        
    # Asegurar que siempre se muestren exactamente 3 bloques según el formato ALM-FORM-006
    while len(activos_list) < 3:
        activos_list.append({
            'codigo': "",
            'descripcion': "",
            'cantidad': "",
            'um': "",
            'marca': "",
            'modelo': "",
            'serial': "",
        })
        
    # Limitar a 3 si por alguna razón hay más (el formato físico solo soporta 3)
    activos_list = activos_list[:3]
    
    # 5. Renderizado del PDF usando la plantilla CORRECTA (ALM-FORM-006)
    template_name = 'inventario/guia_transferencia_pdf.html'
    context = {
        'guia': guia_data,
        'activos': activos_list,
    }
    
    html_string = render_to_string(template_name, context, request=request)
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    pdf = html.write_pdf()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Transferencia_{guia_obj.nro_guia}.pdf"'
    
    return response



# ==================================================
# VISTA 14: Reportes (Listado detallado con botones)
# ==================================================
@login_required(login_url='login')
def reportes(request):
    # 1. Obtener parámetros de Filtro Inteligente Kárdex-Excel
    f_rp = request.GET.get('f_rp', '').strip()
    f_em = request.GET.get('f_em', '').strip()
    f_rq = request.GET.get('f_rq', '').strip()
    f_dpto = request.GET.get('f_dpto', '').strip()
    f_codigo = request.GET.get('f_codigo', '').strip()
    f_desc = request.GET.get('f_desc', '').strip()
    f_np = request.GET.get('f_np', '').strip()
    f_odc = request.GET.get('f_odc', '').strip()
    f_cant_sol = request.GET.get('f_cant_sol', '').strip()
    f_um = request.GET.get('f_um', '').strip()
    f_prov = request.GET.get('f_prov', '').strip()
    f_nota = request.GET.get('f_nota', '').strip()
    f_cant_rec = request.GET.get('f_cant_rec', '').strip()
    f_pu = request.GET.get('f_pu', '').strip()
    f_fecha = request.GET.get('f_fecha', '').strip()
    
    # EXCLUSIÓN DE MIGRACIÓN Y PENDIENTES: Ocultamos saldos iniciales y registros sin ítem asignado
    items_qs = DetalleRecepcion.objects.exclude(
        Q(es_saldo_inicial=True) | 
        (Q(material__isnull=True) & Q(activo__isnull=True) & Q(gasto_directo__isnull=True)) |
        (Q(gasto_directo__isnull=False) & Q(reporte__isnull=True))
    ).select_related('material', 'activo', 'gasto_directo', 'reporte').all().order_by('-fecha_recepcion', '-id')

    # 2. Aplicar Filtro Acumulativo de Precisión
    if f_rp:
        items_qs = items_qs.filter(reporte__nro_reporte__icontains=f_rp)
    if f_em:
        items_qs = items_qs.filter(nro_control_entrada__icontains=f_em)
    if f_rq:
        items_qs = items_qs.filter(nro_rq__icontains=f_rq)
    if f_dpto:
        items_qs = items_qs.filter(departamento__icontains=f_dpto)
    if f_codigo:
        items_qs = items_qs.filter(
            Q(material__codigo__icontains=f_codigo) |
            Q(activo__codigo_activo__icontains=f_codigo) |
            Q(gasto_directo__codigo_dg__icontains=f_codigo)
        )
    if f_desc:
        items_qs = items_qs.filter(
            Q(material__descripcion__icontains=f_desc) |
            Q(activo__descripcion__icontains=f_desc) |
            Q(gasto_directo__descripcion__icontains=f_desc) |
            Q(descripcion_entrada__icontains=f_desc)
        )
    if f_np:
        items_qs = items_qs.filter(material__nro_parte__icontains=f_np)
    if f_odc:
        items_qs = items_qs.filter(nro_odc__icontains=f_odc)
    if f_cant_sol:
        items_qs = items_qs.filter(cantidad_solicitada__icontains=f_cant_sol)
    if f_um:
        items_qs = items_qs.filter(
            Q(material__unidad_medida__icontains=f_um) |
            Q(gasto_directo__unidad_medida_dg__icontains=f_um)
        )
    if f_prov:
        items_qs = items_qs.filter(proveedor__icontains=f_prov)
    if f_nota:
        items_qs = items_qs.filter(nro_nota_entrega__icontains=f_nota)
    if f_cant_rec:
        items_qs = items_qs.filter(cantidad_recibida__icontains=f_cant_rec)
    if f_pu:
        items_qs = items_qs.filter(precio_unitario__icontains=f_pu)
    if f_fecha:
        items_qs = items_qs.filter(fecha_recepcion__icontains=f_fecha)

    # Paginar resultados
    paginator = Paginator(items_qs, 50)
    page_number = request.GET.get('page')
    items_paginados = paginator.get_page(page_number)

    # Auditoría: Conteo de entradas sin material/activo asignado (Monederos)
    total_pendientes = DetalleRecepcion.objects.filter(material__isnull=True, activo__isnull=True, gasto_directo__isnull=True).count()

    # --- LÓGICA CENTRALIZADA: Garantizar que siempre haya uno ABIERTO ---
    if not ReporteRecepcion.objects.filter(estado='ABIERTO').exists():
        ReporteRecepcion.objects.create(estado='ABIERTO')
    
    hay_reportes_abiertos = True

    contexto = {
        'reportes': items_paginados,
        'query': '',
        'total_pendientes': total_pendientes,
        'hay_reportes_abiertos': hay_reportes_abiertos
    }
    return render(request, 'inventario/reportes.html', contexto)

@login_required(login_url='login')
@user_passes_test(es_almacenista, login_url='reportes')
def editar_reporte(request, pk):
    reporte = get_object_or_404(ReporteRecepcion, pk=pk)
    if request.method == 'POST':
        # 1. Validar Firma Digital (Contraseña)
        password_firma = request.POST.get('password_firma', '').strip()
        if not password_firma or not request.user.check_password(password_firma):
            from django.contrib import messages
            messages.error(request, "Firma digital inválida: La contraseña ingresada es incorrecta. Edición cancelada.")
            return redirect('reportes')
            
        form = ReporteRecepcionEditForm(request.POST, instance=reporte)
        if form.is_valid():
            form.save()
            from django.contrib import messages
            messages.success(request, f"Reporte {reporte.nro_reporte} actualizado correctamente.")
            return redirect('reportes')
    else:
        form = ReporteRecepcionEditForm(instance=reporte)
    
    return render(request, 'inventario/editar_reporte.html', {
        'form': form, 
        'reporte': reporte
    })

@login_required(login_url='login')
@user_passes_test(es_almacenista, login_url='reportes')
@login_required(login_url='login')
@roles_requeridos(['Coordinador de Almacén', 'Especialista de Activos'])
def eliminar_reporte(request, pk):
    from decimal import Decimal
    reporte = get_object_or_404(ReporteRecepcion, pk=pk)
    nro_reporte = reporte.nro_reporte
    
    if request.method == 'POST':
        # 1. Validar Firma Digital (Contraseña)
        password_firma = request.POST.get('password_firma', '').strip()
        if not password_firma or not request.user.check_password(password_firma):
            from django.contrib import messages
            messages.error(request, "Firma digital inválida: La contraseña ingresada es incorrecta. Eliminación cancelada.")
            return redirect('reportes')
            
        try:
            with transaction.atomic():
                # A. Obtener todos los detalles de recepción desglosados asociados al reporte
                detalles = DetalleRecepcion.objects.filter(reporte=reporte)
                
                # Para recrear el monedero original, tomamos los datos de cualquiera de los ítems
                primer_detalle = detalles.first()
                
                if primer_detalle:
                    # Calcular el valor total sumado de todos los detalles desglosados
                    total_valor_monedero = Decimal('0.00')
                    for det in detalles:
                        # Revertir el stock en el Maestro de Materiales/Activos
                        if det.material:
                            det.material.stock_actual -= det.cantidad_recibida
                            det.material.save()
                        elif det.activo:
                            det.activo.stock -= int(det.cantidad_recibida)
                            det.activo.save()
                        
                        total_valor_monedero += det.valor_recibido
                    
                    # Recrear el monedero original como entrada global pendiente (sin material)
                    monedero_restaurado = DetalleRecepcion(
                        reporte=None, # Vuelve como solicitud/pendiente de desglose
                        material=None,
                        activo=None,
                        gasto_directo=None,
                        tipo_ingreso=primer_detalle.tipo_ingreso,
                        nro_control_entrada=primer_detalle.nro_control_entrada,
                        nro_rq=primer_detalle.nro_rq,
                        departamento=primer_detalle.departamento,
                        nro_odc=primer_detalle.nro_odc,
                        nro_nota_entrega=primer_detalle.nro_nota_entrega,
                        proveedor=primer_detalle.proveedor,
                        fecha_recepcion=primer_detalle.fecha_recepcion,
                        cantidad_solicitada=Decimal('1.00'),
                        cantidad_recibida=Decimal('1.00'),
                        precio_unitario=total_valor_monedero,
                        moneda=primer_detalle.moneda or 'USD',
                        descripcion_entrada=primer_detalle.descripcion_entrada,
                        observaciones=primer_detalle.observaciones
                    )
                    monedero_restaurado.save()
                
                # B. Eliminar los detalles desglosados individuales
                detalles.delete()
                
                # C. Eliminar el reporte padre
                reporte.delete()
                
                # Registrar actividad de auditoría
                registrar_actividad(request.user, "Eliminó Reporte", f"Eliminó el reporte {nro_reporte} y restauró la entrada original como pendiente.")
                
                from django.contrib import messages
                messages.success(request, f"Reporte {nro_reporte} eliminado con éxito. El stock fue descontado del registro maestro y la entrada original fue restaurada como pendiente de desglose en la bandeja.")
        except Exception as e:
            from django.contrib import messages
            messages.error(request, f"Error al eliminar reporte: {str(e)}")
            
    return redirect('reportes')

# VISTA PARA GENERAR PDF DE RECEPCIONES (OPTIMIZADA)
@login_required(login_url='login')
@roles_requeridos(['Coordinador de Almacén', 'Especialista de Activos'])
def generar_reporte_recepcion_pdf(request):
    from django.template.loader import render_to_string
    from weasyprint import HTML
    
    # 1. Obtener el último Reporte de Recepción (el lote más reciente)
    ultimo_reporte = ReporteRecepcion.objects.order_by('-id').first()
    
    # 2. Filtrar los registros que pertenecen únicamente a ese reporte
    # Mantenemos la exclusión de Saldos Iniciales y registros Pendientes (sin material)
    if ultimo_reporte:
        reportes_qs = DetalleRecepcion.objects.filter(
            reporte=ultimo_reporte
        ).exclude(
            Q(es_saldo_inicial=True) | 
            (Q(material__isnull=True) & Q(activo__isnull=True) & Q(gasto_directo__isnull=True))
        ).select_related('material', 'activo', 'gasto_directo', 'reporte').order_by('id')
    else:
        reportes_qs = DetalleRecepcion.objects.none()
    
    # 3. Relleno de filas vacías para mantener el diseño (mantenemos tope de 20 para el diseño visual)
    count = reportes_qs.count()
    filas_vacias = range(max(0, 20 - count))
    
    context = {
        'reporte_padre': ultimo_reporte,
        'reportes': reportes_qs,
        'filas_vacias': filas_vacias,
        'hoy': dt.date.today(),
    }
    
    # 3. Renderizado y Generación de PDF
    html_string = render_to_string('inventario/reporte_pdf.html', context, request=request)
    html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
    pdf = html.write_pdf()
    
    # 4. Respuesta HTTP
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="Reporte_Recepcion_Materiales.pdf"'
    return response

@login_required(login_url='login')
@roles_requeridos(['Coordinador de Almacén', 'Especialista de Activos'])
def cargar_partidas_csv(request):
    """Carga masiva de catálogo de finanzas (Partidas/Cuentas) desde CSV."""
    import csv
    import io
    from django.db import transaction
    from django.contrib import messages
    from django.utils import timezone
    
    if request.method == 'POST':
        archivo = request.FILES.get('archivo_csv')
        if not archivo:
            messages.error(request, "No se seleccionó ningún archivo.")
            return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
        
        try:
            with transaction.atomic():
                # 1. Decodificación robusta (UTF-8-SIG para quitar el BOM de Excel)
                content = archivo.read()
                try:
                    data_set = content.decode('utf-8-sig')
                except UnicodeDecodeError:
                    data_set = content.decode('latin-1')
                
                io_string = io.StringIO(data_set)
                
                # 2. Detección de delimitador (Soporta , y ;)
                sample = io_string.read(2048)
                io_string.seek(0)
                if not sample:
                    raise Exception("El archivo está vacío.")
                
                dialect = csv.Sniffer().sniff(sample)
                reader = csv.DictReader(io_string, dialect=dialect)
                
                # 3. Auditoría de Encabezados (Debug)
                # Normalizamos los fieldnames para ignorar espacios y mayúsculas/minúsculas
                headers_originales = reader.fieldnames
                headers_limpios = [h.strip().upper() for h in headers_originales if h]
                print(f"DEBUG: Encabezados originales: {headers_originales}")
                print(f"DEBUG: Encabezados procesados: {headers_limpios}")
                
                mandatory_columns = ['DEPARTAMENTO', 'CUENTA_CONTABLE', 'PARTIDA_PRESUPUESTARIA']
                for col in mandatory_columns:
                    if col not in headers_limpios:
                        raise Exception(f"Falta la columna obligatoria: {col}. Detectadas: {headers_limpios}")

                año_actual = timezone.now().year
                count = 0
                
                for row_raw in reader:
                    # Crear un nuevo diccionario con claves normalizadas (Mayúsculas y sin espacios)
                    # Esto resuelve el problema si el CSV viene con " departamento" o "Departamento"
                    row = {k.strip().upper(): v for k, v in row_raw.items() if k}
                    
                    # 4. Extracción y Limpieza Defensiva
                    depto = str(row.get('DEPARTAMENTO', '')).strip()
                    cuenta = str(row.get('CUENTA_CONTABLE', '')).strip()
                    partida = str(row.get('PARTIDA_PRESUPUESTARIA', '')).strip()
                    
                    # Validación: No permitir registros vacíos en campos clave
                    if not depto or not cuenta or not partida:
                        print(f"⚠️ Fila saltada por datos incompletos: Depto={depto}, Cta={cuenta}, Partida={partida}")
                        continue
                    
                    # 5. Persistencia en base de datos
                    PresupuestoAnual.objects.update_or_create(
                        anio=año_actual,
                        departamento=depto,
                        cuenta_contable=cuenta,
                        partida=partida,
                        defaults={
                            'descripcion_cuenta': str(row.get('DESCRIPCION_CUENTA', '')).strip(),
                            'rubro_1': str(row.get('RUBRO_1', '')).strip(),
                            'rubro_2': str(row.get('RUBRO_2', '')).strip(),
                        }
                    )
                    count += 1
                
                if count == 0:
                    raise Exception("No se encontraron filas válidas para procesar. Verifique el formato de su CSV.")
                    
                messages.success(request, f"¡Migración exitosa! Se procesaron {count} reglas financieras correctamente.")
                
        except Exception as e:
            error_msg = f"Error al procesar CSV: {str(e)}"
            messages.error(request, error_msg)
            print(f"❌ ERROR CARGA FINANZAS: {str(e)}")
            
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))
# ==========================================
# NUEVA VISTA: BANDEJA DE ENTRADA DEL JEFE
# ==========================================
@login_required(login_url='login')
@roles_requeridos(['Coordinador de Almacén', 'Especialista de Activos'])
def reportes_pendientes(request):
    # Traemos SOLO los registros globales (Monederos) que no han sido desglosados
    pendientes = DetalleRecepcion.objects.filter(
        material__isnull=True,
        activo__isnull=True,
        gasto_directo__isnull=True
    ).select_related('reporte').order_by('fecha_recepcion', '-id')

    # Formularios para el desglose (se usarán vía JS en la misma página)
    form = ReporteRecepcionForm()
    form_detalle = DetalleRecepcionForm()
    
    # Lista de materiales y activos para el Select2
    materiales = Material.objects.all().order_by('codigo')
    activos = Activo.objects.all().order_by('codigo_activo')
    centros_costo = CentroCosto.objects.all().order_by('nombre')

    contexto = {
        'pendientes': pendientes,
        'total_pendientes': pendientes.count(),
        'form': form,
        'form_detalle': form_detalle,
        'materiales': materiales,
        'activos': activos,
        'centros_costo': centros_costo
    }
    return render(request, 'inventario/reportes_pendientes.html', contexto)
# ==================================================
# API: Obtener datos de Material por AJAX
# ==================================================
@login_required(login_url='login')
def get_material_info(request, material_id):
    from django.http import JsonResponse
    try:
        material = Material.objects.get(id=material_id)
        data = {
            'descripcion': material.descripcion,
            'nro_parte': material.nro_parte or 'N/A',
            'unidad_medida': material.unidad_medida,
            'cargo': material.cargo
        }
        return JsonResponse(data)
    except Material.DoesNotExist:
        return JsonResponse({'error': 'Material no encontrado'}, status=404)


# ==================================================
# API: Obtener desglose de LOTES (FIFO) de un Material
# ==================================================
@login_required(login_url='login')
def api_lotes_material(request, material_id):
    from django.http import JsonResponse
    material = get_object_or_404(Material, id=material_id)
    
    # Obtenemos todos los lotes que tienen stock disponible
    lotes_qs = material.entradas.filter(
        cantidad_recibida__gt=0
    ).order_by('fecha_recepcion', 'id')
    
    lotes = []
    for lote in lotes_qs:
        # Solo incluimos lotes con disponibilidad real
        if lote.cantidad_disponible > 0:
            lotes.append({
                'em': lote.nro_control_entrada,
                'fecha': lote.fecha_recepcion.strftime('%d/%m/%Y'),
                'odc': lote.nro_odc,
                'recibido': float(lote.cantidad_recibida),
                'disponible': float(lote.cantidad_disponible),
                'precio': float(lote.precio_unitario or 0),
                'total': float(lote.valor_recibido),
            })
    
    return JsonResponse({
        'codigo': material.codigo,
        'descripcion': material.descripcion,
        'stock_total': float(material.stock_actual),
        'lotes': lotes
    })


# ==================================================
# API: Obtener desglose de LOTES (FIFO) de un Activo
# ==================================================
@login_required(login_url='login')
def api_lotes_activo(request, activo_id):
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404
    from inventario.models import Activo
    
    activo = get_object_or_404(Activo, id=activo_id)
    
    # Obtenemos todos los lotes que tienen stock disponible
    lotes_qs = activo.entradas.filter(
        cantidad_recibida__gt=0
    ).order_by('fecha_recepcion', 'id')
    
    lotes = []
    for lote in lotes_qs:
        if lote.cantidad_disponible > 0:
            lotes.append({
                'em': lote.nro_control_entrada,
                'fecha': lote.fecha_recepcion.strftime('%d/%m/%Y'),
                'odc': lote.nro_odc,
                'recibido': float(lote.cantidad_recibida),
                'disponible': float(lote.cantidad_disponible),
                'precio': float(lote.precio_unitario or 0),
                'total': float(lote.valor_recibido),
            })
            
    return JsonResponse({
        'codigo': activo.codigo_activo,
        'descripcion': activo.descripcion,
        'stock_total': float(activo.stock_final),
        'lotes': lotes
    })


# ==================================================
# API: Obtener partidas presupuestarias por departamento (AJAX)
# ==================================================
@login_required(login_url='login')
def api_partidas_por_departamento(request):
    from django.http import JsonResponse
    import datetime
    
    departamento = request.GET.get('departamento', '').strip()
    anio = datetime.date.today().year
    
    if not departamento:
        return JsonResponse([], safe=False)
    
    partidas = PresupuestoAnual.objects.filter(
        departamento__iexact=departamento,
        anio=anio
    ).values('id', 'partida', 'cuenta_contable', 'descripcion_cuenta', 'rubro_1', 'rubro_2')
    
    return JsonResponse(list(partidas), safe=False)


# ==================================================
# API: Historial de una ODC + Nota de Entrega (Para el panel lateral en Entradas)
# Devuelve todos los DetalleRecepcion con la misma ODC y Nota de Entrega,
# tanto entradas simples como ítems de reportes.
# ==================================================
@login_required(login_url='login')
def api_historial_odc(request):
    from django.http import JsonResponse

    odc = request.GET.get('odc', '').strip()
    nota = request.GET.get('nota', '').strip()

    if not odc:
        return JsonResponse({'entradas': [], 'reportes': []})

    # Filtrar por ODC (obligatorio) y nota de entrega si viene
    filtro = Q(nro_odc=odc)
    if nota:
        filtro &= Q(nro_nota_entrega=nota)

    registros = DetalleRecepcion.objects.filter(filtro).select_related('material', 'reporte').order_by('fecha_recepcion', 'id')

    entradas = []
    reportes_vistos = set()
    reportes_lista = []

    for r in registros:
        desc = r.descripcion_entrada or (r.material.descripcion if r.material else '-')
        em = r.nro_control_entrada or '-'
        entradas.append({
            'em': em,
            'fecha': r.fecha_recepcion.strftime('%d/%m/%Y') if r.fecha_recepcion else '-',
            'odc': r.nro_odc or '-',
            'nota': r.nro_nota_entrega or '-',
            'proveedor': r.proveedor or '-',
            'descripcion': desc,
            'costo': str(r.precio_unitario or '0.00'),
            'reporte': r.reporte.nro_reporte if r.reporte else None,
        })
        if r.reporte and r.reporte.id not in reportes_vistos:
            reportes_vistos.add(r.reporte.id)
            reportes_lista.append({
                'nro_reporte': r.reporte.nro_reporte,
                'fecha': r.reporte.fecha_recepcion.strftime('%d/%m/%Y') if r.reporte.fecha_recepcion else '-',
            })

    return JsonResponse({'entradas': entradas, 'reportes': reportes_lista})

@login_required(login_url='login')
@roles_requeridos(['Coordinador de Almacén', 'Especialista de Activos'])
def desglosar_entrada(request, detalle_id):
    monedero = get_object_or_404(DetalleRecepcion, id=detalle_id)
    
    if request.method == 'POST':
        # 1. Validar Firma Digital (Contraseña)
        password_firma = request.POST.get('password_firma', '').strip()
        if not password_firma or not request.user.check_password(password_firma):
            from django.contrib import messages
            messages.error(request, "Firma digital inválida: La contraseña ingresada es incorrecta. Transacción cancelada.")
            return redirect('reportes_pendientes')

        carrito_json = request.POST.get('carrito_datos', '[]')
        try:
            items_carrito = json.loads(carrito_json)
        except:
            items_carrito = []

        # Detección de ítems por array directo (EDG)
        codigos_edg = request.POST.getlist('codigo_edg[]')

        if items_carrito or codigos_edg:
            with transaction.atomic():
                reporte_obj = ReporteRecepcion.objects.filter(estado='ABIERTO').first()
                if not reporte_obj:
                    reporte_obj = ReporteRecepcion.objects.create(estado='ABIERTO')

                for item in items_carrito:
                    es_edg_item = item.get('es_edg') is True or str(item.get('es_edg')).lower() == 'true'
                    material_obj = None
                    activo_obj = None
                    gasto_directo_obj = None
                    tipo_ingreso = 'Material'

                    if es_edg_item:
                        tipo_ingreso = 'Directo al Gasto'
                        cod_edg = item.get('codigo_edg', '').strip().upper()
                        desc_edg = item.get('descripcion_edg', '').strip().upper()
                        um_edg = item.get('um_edg', '').strip().upper()
                        cargo_edg = item.get('cargo_edg', 'OPERACIONES')

                        if cod_edg and desc_edg:
                            gd_obj, _ = GastoDirecto.objects.get_or_create(
                                codigo_dg=cod_edg,
                                defaults={
                                    'descripcion': desc_edg,
                                    'unidad_medida_dg': um_edg,
                                    'cargo_dg': cargo_edg,
                                    'stock': 0
                                }
                            )
                            gasto_directo_obj = gd_obj
                    else:
                        codigo_material = item.get('material', '').split(' - ')[0].replace('[MATERIAL]', '').replace('[ACTIVOS]', '').replace('[DIRECTO AL GASTO]', '').strip()
                        material_obj = Material.objects.filter(codigo=codigo_material).first()
                        if not material_obj:
                            activo_obj = Activo.objects.filter(codigo_activo=codigo_material).first()
                        tipo_ingreso = 'Material' if material_obj else 'Activo'

                    if not material_obj and not activo_obj and not gasto_directo_obj: 
                        continue

                    nuevo_detalle = DetalleRecepcion(
                        reporte=reporte_obj,
                        material=material_obj,
                        activo=activo_obj,
                        gasto_directo=gasto_directo_obj,
                        tipo_ingreso=tipo_ingreso,
                        nro_control_entrada=monedero.nro_control_entrada,
                        nro_rq=item.get('nro_rq') or monedero.nro_rq,
                        departamento=item.get('departamento') or monedero.departamento,
                        nro_odc=item.get('nro_odc') or monedero.nro_odc,
                        nro_nota_entrega=item.get('nro_nota_entrega') or monedero.nro_nota_entrega,
                        proveedor=item.get('proveedor') or monedero.proveedor,
                        fecha_recepcion=monedero.fecha_recepcion,
                        cantidad_solicitada=Decimal(item.get('cantidad_solicitada') or '0'),
                        cantidad_recibida=Decimal(item.get('cantidad_recibida') or '0'),
                        precio_unitario=Decimal(item.get('precio_unitario') or '0'),
                        moneda=monedero.moneda or 'USD',
                        descripcion_entrada=monedero.descripcion_entrada,
                        observaciones=monedero.observaciones
                    )
                    nuevo_detalle.save()

                # B. Procesar ítems desde Arrays Directos (EDG)
                if codigos_edg:
                    descripciones_edg = request.POST.getlist('descripcion_edg[]')
                    ums_edg = request.POST.getlist('um_edg[]')
                    cargos_edg = request.POST.getlist('cargo_edg[]')
                    cantidades_edg = request.POST.getlist('cantidad_edg[]')
                    precios_edg = request.POST.getlist('precio_edg[]')
                    odcs_edg = request.POST.getlist('odc_edg[]')
                    notas_edg = request.POST.getlist('nota_edg[]')

                    for cod, desc, um, cargo, cant, precio, odc, nota in zip(
                        codigos_edg, descripciones_edg, ums_edg, cargos_edg, 
                        cantidades_edg, precios_edg, odcs_edg, notas_edg
                    ):
                        if not cod or not desc: continue

                        gd_obj, _ = GastoDirecto.objects.get_or_create(
                            codigo_dg=cod.strip().upper(),
                            defaults={
                                'descripcion': desc.strip().upper(),
                                'unidad_medida_dg': um.strip().upper(),
                                'cargo_dg': cargo.strip().upper(),
                                'stock': 0
                            }
                        )

                        nuevo_detalle = DetalleRecepcion(
                            reporte=reporte_obj,
                            gasto_directo=gd_obj,
                            tipo_ingreso='Directo al Gasto',
                            nro_control_entrada=monedero.nro_control_entrada,
                            nro_rq=monedero.nro_rq,
                            departamento=monedero.departamento,
                            nro_odc=odc.strip().upper() or monedero.nro_odc,
                            nro_nota_entrega=nota.strip().upper() or monedero.nro_nota_entrega,
                            proveedor=monedero.proveedor,
                            fecha_recepcion=monedero.fecha_recepcion,
                            cantidad_recibida=Decimal(cant or '0'),
                            precio_unitario=Decimal(precio or '0'),
                            moneda=monedero.moneda or 'USD',
                            descripcion_entrada=monedero.descripcion_entrada,
                            observaciones=monedero.observaciones
                        )
                        nuevo_detalle._tipo_entrada_manual = 'DIRECTO AL GASTO'
                        nuevo_detalle.save()

                # Registrar actividad de auditoría
                nro_ems = monedero.nro_control_entrada if monedero.nro_control_entrada else 'N/A'
                registrar_actividad(request.user, "Desglosó Entrada", f"Desglose de entrada de material para el monedero/control {nro_ems}")

                # C. Finalización: Eliminar el monedero original una vez desglosado
                monedero.delete()
            return redirect('reportes_pendientes')

@login_required(login_url='login')
@roles_requeridos(['Coordinador de Almacén', 'Especialista de Activos'])
def cambiar_estado_reportes(request):
    if request.method == 'POST':
        from django.http import JsonResponse
        import json
        try:
            data = json.loads(request.body)
            nuevo_estado = data.get('estado')
            if nuevo_estado == 'CERRADO':
                with transaction.atomic():
                    # 1. Buscamos el reporte que está actualmente abierto
                    reporte_abierto = ReporteRecepcion.objects.filter(estado='ABIERTO').first()
                    if reporte_abierto:
                        reporte_abierto.estado = 'CERRADO'
                        reporte_abierto.save()
                    
                    # 2. Creamos el nuevo reporte para el siguiente ciclo
                    ReporteRecepcion.objects.create(estado='ABIERTO')
                
                return JsonResponse({'status': 'ok', 'mensaje': 'Reporte cerrado y nuevo reporte abierto.'})
            
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error'}, status=400)

@login_required(login_url='login')
@user_passes_test(es_almacenista, login_url='reportes')
def actualizar_ubicacion_material(request):
    if request.method == 'POST':
        import json
        from django.http import JsonResponse
        try:
            data = json.loads(request.body)
            material_id = data.get('material_id')
            nueva_ubicacion = data.get('ubicacion')
            
            if material_id:
                material = get_object_or_404(Material, id=material_id)
                material.ubicacion = nueva_ubicacion
                material.save()
                return JsonResponse({'status': 'ok'})
            
            return JsonResponse({'status': 'error', 'message': 'Falta ID de Material'}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error'}, status=400)

@login_required(login_url='login')
@user_passes_test(es_almacenista, login_url='reportes')
def actualizar_volumen_carpeta(request):
    if request.method == 'POST':
        import json
        from django.http import JsonResponse
        try:
            data = json.loads(request.body)
            nro_em = data.get('nro_control_entrada')
            nuevo_volumen = data.get('volumen')
            
            if nro_em:
                # Actualizamos todos los registros que compartan ese EM (para el caso de desgloses)
                DetalleRecepcion.objects.filter(nro_control_entrada=nro_em).update(volumen_carpeta=nuevo_volumen)
                return JsonResponse({'status': 'ok'})
            
            return JsonResponse({'status': 'error', 'message': 'Falta Nro. Control'}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error'}, status=400)


# ==================================================
# VISTAS DE CONSUMO ANUAL (WEB & EXCEL)
# ==================================================

@login_required(login_url='login')
def consumo_anual_vista(request):
    """Vista web para visualizar el historial de auditoría contable con filtros y paginación."""
    
    # 1. Obtener parámetros de filtro (Nombres estandarizados)
    f_centro = request.GET.get('centro_costo', '')
    f_desde = request.GET.get('fecha_desde', '')
    f_hasta = request.GET.get('fecha_hasta', '')

    # Filtros inteligentes por columna (Kárdex-Excel)
    f_codigo = request.GET.get('f_codigo', '').strip()
    f_desc = request.GET.get('f_desc', '').strip()
    f_np = request.GET.get('f_np', '').strip()
    f_odc = request.GET.get('f_odc', '').strip()
    f_cant = request.GET.get('f_cant', '').strip()
    f_um = request.GET.get('f_um', '').strip()
    f_pu = request.GET.get('f_pu', '').strip()
    f_monto = request.GET.get('f_monto', '').strip()
    f_rim = request.GET.get('f_rim', '').strip()
    f_centro_text = request.GET.get('f_centro_text', '').strip()
    f_sm = request.GET.get('f_sm', '').strip()
    f_cuenta = request.GET.get('f_cuenta', '').strip()
    f_desc_cuenta = request.GET.get('f_desc_cuenta', '').strip()
    f_partida = request.GET.get('f_partida', '').strip()
    f_rubro1 = request.GET.get('f_rubro1', '').strip()
    f_rubro2 = request.GET.get('f_rubro2', '').strip()
    f_fecha = request.GET.get('f_fecha', '').strip()

    # 2. Queryset base (Excluimos ajustes de migración)
    despachos_list = SalidaMaterialDetalle.objects.exclude(
        Q(salida__nro_rim__startswith='AJUSTE-MIG') | Q(salida__departamento='MIGRACIÓN')
    ).select_related(
        'salida__material', 
        'salida__activo',
        'salida__centro_costo', 
        'detalle_recepcion'
    )

    # 3. Filtrado Dinámico (Búsqueda Híbrida para integridad histórica)
    if f_centro:
        try:
            centro_obj = CentroCosto.objects.get(id=f_centro)
            despachos_list = despachos_list.filter(
                Q(salida__centro_costo_id=f_centro) | 
                Q(salida__centro_costo_texto=centro_obj.nombre)
            )
        except CentroCosto.DoesNotExist:
            despachos_list = despachos_list.filter(salida__centro_costo_id=f_centro)
    
    if f_desde:
        despachos_list = despachos_list.filter(salida__fecha_despacho__gte=f_desde)
    
    if f_hasta:
        despachos_list = despachos_list.filter(salida__fecha_despacho__lte=f_hasta)

    # Aplicar Filtros Inteligentes por Columna
    if f_codigo:
        despachos_list = despachos_list.filter(
            Q(salida__material__codigo__icontains=f_codigo) |
            Q(salida__activo__codigo_activo__icontains=f_codigo)
        )
    if f_desc:
        despachos_list = despachos_list.filter(
            Q(salida__material__descripcion__icontains=f_desc) |
            Q(salida__activo__descripcion__icontains=f_desc)
        )
    if f_np:
        despachos_list = despachos_list.filter(
            Q(salida__material__nro_parte__icontains=f_np) |
            Q(salida__activo__nro_parte_ac__icontains=f_np)
        )
    if f_odc:
        despachos_list = despachos_list.filter(detalle_recepcion__nro_odc__icontains=f_odc)
    if f_cant:
        despachos_list = despachos_list.filter(cantidad__icontains=f_cant)
    if f_um:
        despachos_list = despachos_list.filter(
            Q(salida__material__unidad_medida__icontains=f_um) |
            Q(salida__activo__unidad_medida_ac__icontains=f_um)
        )
    if f_pu:
        despachos_list = despachos_list.filter(precio_unitario__icontains=f_pu)
    if f_monto:
        despachos_list = despachos_list.filter(subtotal__icontains=f_monto)
    if f_rim:
        despachos_list = despachos_list.filter(salida__nro_rim__icontains=f_rim)
    if f_centro_text:
        despachos_list = despachos_list.filter(
            Q(salida__centro_costo__nombre__icontains=f_centro_text) |
            Q(salida__centro_costo_texto__icontains=f_centro_text)
        )
    if f_sm:
        despachos_list = despachos_list.filter(salida__nro_sm__icontains=f_sm)
    if f_cuenta:
        despachos_list = despachos_list.filter(salida__cuenta_contable__icontains=f_cuenta)
    if f_desc_cuenta:
        despachos_list = despachos_list.filter(salida__descripcion_cuenta__icontains=f_desc_cuenta)
    if f_partida:
        despachos_list = despachos_list.filter(salida__partida_presupuestaria__icontains=f_partida)
    if f_rubro1:
        despachos_list = despachos_list.filter(salida__rubro_1__icontains=f_rubro1)
    if f_rubro2:
        despachos_list = despachos_list.filter(salida__rubro_2__icontains=f_rubro2)
    if f_fecha:
        despachos_list = despachos_list.filter(salida__fecha_despacho__icontains=f_fecha)

    # Ordenar por fecha descendente
    despachos_list = despachos_list.order_by('-salida__fecha_despacho', '-id')

    # 4. Datos para el formulario (Centros de Costo)
    centros = CentroCosto.objects.all().order_by('nombre')

    # 5. Paginación: 50 registros por página
    paginator = Paginator(despachos_list, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 6. Preservar estado
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    query_prefix = query_params.urlencode() + '&' if query_params else ''

    contexto = {
        'despachos': page_obj,
        'centros': centros,
        'query_prefix': query_prefix,
        'filtros': {
            'centro_costo': f_centro,
            'fecha_desde': f_desde,
            'fecha_hasta': f_hasta,
            'codigo': f_codigo,
            'desc': f_desc,
            'np': f_np,
            'odc': f_odc,
            'cant': f_cant,
            'um': f_um,
            'pu': f_pu,
            'monto': f_monto,
            'rim': f_rim,
            'centro_text': f_centro_text,
            'sm': f_sm,
            'cuenta': f_cuenta,
            'desc_cuenta': f_desc_cuenta,
            'partida': f_partida,
            'rubro1': f_rubro1,
            'rubro2': f_rubro2,
            'fecha': f_fecha,
        }
    }
    return render(request, 'inventario/consumo_anual.html', contexto)


@login_required(login_url='login')
def exportar_consumo_anual_excel(request):
    """Genera y descarga el reporte de Consumo Anual en formato Excel .xlsx"""
    
    # 1. Crear el libro y la hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Consumo Anual Finanzas"

    # 2. Definir Estilos
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    alignment_center = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # 3. Encabezados (18 columnas exactas)
    headers = [
        'ÍTEM', 'CODIGO MATERIAL', 'DESCRIPCION MATERIAL', 'N/P', 
        'ORDEN DE COMPRA', 'CANT. DESPACHADA', 'UNIDAD MEDIDA', 
        'PRECIO UNIT. $', 'MONTO $', 'RIM', 'CENTRO DE COSTO', 'SM', 
        'CUENTA CONTABLE', 'DESCRIPCION DE LA CUENTA CONTABLE', 
        'PARTIDA PRESUPUESTARIA', 'RUBRO 1', 'RUBRO 2', 'FECHA DE ENTREGA'
    ]

    ws.append(headers)

    # Aplicar estilos al encabezado
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = border_style

    # 4. Consultar Datos con Filtros (Captura idéntica)
    f_centro = request.GET.get('centro_costo', '')
    f_desde = request.GET.get('fecha_desde', '')
    f_hasta = request.GET.get('fecha_hasta', '')

    # Filtros inteligentes por columna (Kárdex-Excel)
    f_codigo = request.GET.get('f_codigo', '').strip()
    f_desc = request.GET.get('f_desc', '').strip()
    f_np = request.GET.get('f_np', '').strip()
    f_odc = request.GET.get('f_odc', '').strip()
    f_cant = request.GET.get('f_cant', '').strip()
    f_um = request.GET.get('f_um', '').strip()
    f_pu = request.GET.get('f_pu', '').strip()
    f_monto = request.GET.get('f_monto', '').strip()
    f_rim = request.GET.get('f_rim', '').strip()
    f_centro_text = request.GET.get('f_centro_text', '').strip()
    f_sm = request.GET.get('f_sm', '').strip()
    f_cuenta = request.GET.get('f_cuenta', '').strip()
    f_desc_cuenta = request.GET.get('f_desc_cuenta', '').strip()
    f_partida = request.GET.get('f_partida', '').strip()
    f_rubro1 = request.GET.get('f_rubro1', '').strip()
    f_rubro2 = request.GET.get('f_rubro2', '').strip()
    f_fecha = request.GET.get('f_fecha', '').strip()

    despachos = SalidaMaterialDetalle.objects.exclude(
        Q(salida__nro_rim__startswith='AJUSTE-MIG') | Q(salida__departamento='MIGRACIÓN')
    ).select_related(
        'salida__material', 
        'salida__activo',
        'salida__centro_costo', 
        'detalle_recepcion'
    )

    if f_centro:
        try:
            centro_obj = CentroCosto.objects.get(id=f_centro)
            despachos = despachos.filter(
                Q(salida__centro_costo_id=f_centro) | 
                Q(salida__centro_costo_texto=centro_obj.nombre)
            )
        except CentroCosto.DoesNotExist:
            despachos = despachos.filter(salida__centro_costo_id=f_centro)
    if f_desde:
        despachos = despachos.filter(salida__fecha_despacho__gte=f_desde)
    if f_hasta:
        despachos = despachos.filter(salida__fecha_despacho__lte=f_hasta)

    # Aplicar Filtros Inteligentes por Columna
    if f_codigo:
        despachos = despachos.filter(
            Q(salida__material__codigo__icontains=f_codigo) |
            Q(salida__activo__codigo_activo__icontains=f_codigo)
        )
    if f_desc:
        despachos = despachos.filter(
            Q(salida__material__descripcion__icontains=f_desc) |
            Q(salida__activo__descripcion__icontains=f_desc)
        )
    if f_np:
        despachos = despachos.filter(
            Q(salida__material__nro_parte__icontains=f_np) |
            Q(salida__activo__nro_parte_ac__icontains=f_np)
        )
    if f_odc:
        despachos = despachos.filter(detalle_recepcion__nro_odc__icontains=f_odc)
    if f_cant:
        despachos = despachos.filter(cantidad__icontains=f_cant)
    if f_um:
        despachos = despachos.filter(
            Q(salida__material__unidad_medida__icontains=f_um) |
            Q(salida__activo__unidad_medida_ac__icontains=f_um)
        )
    if f_pu:
        despachos = despachos.filter(precio_unitario__icontains=f_pu)
    if f_monto:
        despachos = despachos.filter(subtotal__icontains=f_monto)
    if f_rim:
        despachos = despachos.filter(salida__nro_rim__icontains=f_rim)
    if f_centro_text:
        despachos = despachos.filter(
            Q(salida__centro_costo__nombre__icontains=f_centro_text) |
            Q(salida__centro_costo_texto__icontains=f_centro_text)
        )
    if f_sm:
        despachos = despachos.filter(salida__nro_sm__icontains=f_sm)
    if f_cuenta:
        despachos = despachos.filter(salida__cuenta_contable__icontains=f_cuenta)
    if f_desc_cuenta:
        despachos = despachos.filter(salida__descripcion_cuenta__icontains=f_desc_cuenta)
    if f_partida:
        despachos = despachos.filter(salida__partida_presupuestaria__icontains=f_partida)
    if f_rubro1:
        despachos = despachos.filter(salida__rubro_1__icontains=f_rubro1)
    if f_rubro2:
        despachos = despachos.filter(salida__rubro_2__icontains=f_rubro2)
    if f_fecha:
        despachos = despachos.filter(salida__fecha_despacho__icontains=f_fecha)

    despachos = despachos.order_by('-salida__fecha_despacho', '-id')

    # 5. Escribir Datos
    for index, d in enumerate(despachos, start=1):
        # Cálculos y valores seguros
        cant = float(d.cantidad)
        precio = float(d.precio_unitario or 0)
        monto = cant * precio
        
        # Lógica híbrida para Activos y Materiales
        if d.salida.activo:
            codigo = d.salida.activo.codigo_activo
            descripcion = d.salida.activo.descripcion
            nro_parte = d.salida.activo.nro_parte_ac or "N/A"
            um = d.salida.activo.unidad_medida_ac
        elif d.salida.material:
            codigo = d.salida.material.codigo
            descripcion = d.salida.material.descripcion
            nro_parte = d.salida.material.nro_parte or "N/A"
            um = d.salida.material.unidad_medida
        else:
            codigo, descripcion, nro_parte, um = "S/N", "ARTICULO NO ENCONTRADO", "N/A", "N/A"

        row = [
            index,                                          # 1. ITEM
            codigo,                                         # 2. CODIGO
            descripcion,                                    # 3. DESCRIPCION
            nro_parte,                                      # 4. N/P
            d.detalle_recepcion.nro_odc,                    # 5. ODC
            cant,                                           # 6. CANTIDAD (Número)
            um,                                             # 7. U.M.
            precio,                                         # 8. PRECIO (Número)
            monto,                                          # 9. MONTO (Número)
            d.salida.nro_rim,                               # 10. RIM
            str(d.salida.centro_costo or d.salida.centro_costo_texto or "-"), # 11. CC
            d.salida.nro_sm or "-",                         # 12. SM
            d.salida.cuenta_contable or "-",                # 13. CUENTA
            d.salida.descripcion_cuenta or "-",             # 14. DESC CUENTA
            d.salida.partida_presupuestaria or "-",         # 15. PARTIDA
            d.salida.rubro_1 or "-",                        # 16. RUBRO 1
            d.salida.rubro_2 or "-",                        # 17. RUBRO 2
            d.salida.fecha_despacho.strftime('%d/%m/%Y')    # 18. FECHA
        ]
        ws.append(row)

    # 6. Auto-ajuste de columnas (Anchos generosos)
    column_widths = [8, 20, 50, 15, 20, 18, 10, 15, 15, 15, 25, 15, 18, 40, 22, 15, 15, 18]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # 7. Preparar Respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Consumo_Anual_{dt.date.today().year}.xlsx"'
    
    wb.save(response)
    return response

@login_required(login_url='login')
def exportar_inventario_maestro_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from decimal import Decimal
    import datetime

    tipo_item = request.GET.get('tipo_item', 'material')
    departamento = request.GET.get('departamento', '').strip()

    wb = Workbook()
    ws = wb.active
    ws.title = f'Valuación {tipo_item.capitalize()}'

    # --- 1. ENCABEZADO CORPORATIVO ---
    # Título (Fila 1)
    titulo = f"REPORTE DE VALORACIÓN DE INVENTARIO - {tipo_item.upper()}"
    ws.merge_cells('A1:G1')
    ws['A1'] = titulo
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Metadatos (Filas 2 y 3)
    fecha_hoy = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ws['A2'] = "Fecha de Emisión:"
    ws['B2'] = fecha_hoy
    ws['A2'].font = Font(bold=True)

    dpto_texto = departamento if departamento else "Todos los Departamentos"
    ws['A3'] = "Departamento Filtrado:"
    ws['B3'] = dpto_texto
    ws['A3'].font = Font(bold=True)

    # Fila 4 vacía (respiro visual)

    # --- 2. ENCABEZADOS DE COLUMNA (Fila 5) ---
    headers = ['UBICACIÓN', 'CÓDIGO', 'DESCRIPCIÓN', 'N/P', 'U/M', 'STOCK', 'VALOR TOTAL ($)']
    ws.append([]) # Fila 4
    ws.append(headers) # Fila 5

    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid') # Gris Claro
    header_font = Font(bold=True)
    centered_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Estilos para ítems agotados
    agotado_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    agotado_font = Font(color='9C0006') # Rojo oscuro (tipo alerta)
    separador_font = Font(bold=True, color='666666', italic=True)

    for cell in ws[5]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.border = border

    gran_total = Decimal('0.00')

    # --- 3. EXTRACCIÓN Y PROCESAMIENTO DE DATOS ---
    if tipo_item == 'material':
        base_qs = Material.objects.exclude(tipo='DIRECTO AL GASTO')
        if departamento:
            base_qs = base_qs.filter(cargo=departamento)
        
        # Dividir para priorizar existencias reales
        items_con_stock = base_qs.filter(stock_actual__gt=0).order_by('codigo')
        items_sin_stock = base_qs.filter(stock_actual__lte=0).order_by('codigo')
        
        # Unir para iteración única con bandera de cambio
        procesar_lista = [ (item, False) for item in items_con_stock ]
        if items_sin_stock.exists():
            procesar_lista.append((None, True)) # Marca de separador
            for item in items_sin_stock:
                procesar_lista.append((item, False))
        
        for item, es_separador in procesar_lista:
            if es_separador:
                ws.append(['', '', '--- ÍTEMS AGOTADOS (STOCK 0) ---', '', '', '', ''])
                row_idx = ws.max_row
                ws.cell(row=row_idx, column=3).font = separador_font
                continue

            valor_item = item.valor_total_inventario
            gran_total += valor_item
            
            ws.append([
                item.ubicacion or '-',
                item.codigo,
                item.descripcion,
                item.nro_parte or '-',
                item.unidad_medida,
                float(item.stock_actual),
                float(valor_item)
            ])

            if item.stock_actual <= 0:
                for cell in ws[ws.max_row]:
                    cell.fill = agotado_fill
                    cell.font = agotado_font

    else:
        # Para Activos, usamos stock_final (property)
        base_qs = Activo.objects.all()
        if departamento:
            base_qs = base_qs.filter(cargo_ac=departamento)
        
        # Convertir a lista y ordenar en Python por stock y código
        activos_list = list(base_qs)
        activos_list.sort(key=lambda x: (x.stock_final == 0, x.codigo_activo))
        
        separador_puesto = False
        for item in activos_list:
            stock = item.stock_final
            
            # Insertar separador visual si pasamos a los de stock 0
            if stock == 0 and not separador_puesto:
                ws.append(['', '', '--- ÍTEMS AGOTADOS (STOCK 0) ---', '', '', '', ''])
                ws.cell(row=ws.max_row, column=3).font = separador_font
                separador_puesto = True

            ultima_entrada = item.entradas.order_by('-fecha_recepcion', '-id').first()
            precio = ultima_entrada.precio_unitario if ultima_entrada else Decimal('0.00')
            valor_item = Decimal(stock) * precio
            gran_total += valor_item
            
            ws.append([
                item.ubicacion or '-',
                item.codigo_activo,
                item.descripcion,
                item.nro_parte_ac or '-',
                item.unidad_medida_ac,
                float(stock),
                float(valor_item)
            ])

            if stock == 0:
                for cell in ws[ws.max_row]:
                    cell.fill = agotado_fill
                    cell.font = agotado_font

    # --- 4. TOTALES Y AJUSTES FINALES ---
    total_row_idx = ws.max_row + 1
    ws.append(['', '', 'SUMA TOTAL DEL INVENTARIO', '', '', '', float(gran_total)])
    
    # Estilo fila total
    ws.cell(row=total_row_idx, column=3).font = Font(bold=True)
    ws.cell(row=total_row_idx, column=7).font = Font(bold=True)
    ws.cell(row=total_row_idx, column=7).number_format = '#,##0.00'

    # Ajustar anchos de columna automáticamente
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    
    for col, value in dims.items():
        # Damos un margen extra y limitamos la descripción
        width = value + 2
        if col == 'C': # Columna Descripción
            width = min(width, 50)
        ws.column_dimensions[col].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Reporte_Valoracion_{tipo_item}_{datetime.date.today()}.xlsx'
    wb.save(response)
    return response


@login_required(login_url='login')
def exportar_entradas_excel(request):
    """
    Genera un reporte Excel detallado de todas las entradas (EM/EA/EDG)
    filtradas por rango de fecha_recepcion.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from decimal import Decimal
    import datetime

    # 1. Captura de Filtros
    f_desde = request.GET.get('fecha_desde', '')
    f_hasta = request.GET.get('fecha_hasta', '')

    # 2. QuerySet Base
    entradas = DetalleRecepcion.objects.select_related('material', 'activo', 'gasto_directo', 'reporte').all()

    if f_desde:
        entradas = entradas.filter(fecha_recepcion__gte=f_desde)
    if f_hasta:
        entradas = entradas.filter(fecha_recepcion__lte=f_hasta)

    entradas = entradas.order_by('-fecha_recepcion', '-id')

    # 3. Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Control General de Entradas"

    # Encabezados Estrictos
    headers = [
        'No. Requisicion', 'CODIGO MATERIAL', 'Descripción Del Material', 'Numero de Parte', 
        'Cargo', 'No. Orden de Compra', 'Fecha Firma ODC', 'Proveedor', 'Moneda', 
        'Cant. Solicitada', 'U/M', 'U.P.', 'Total', 'ETA', 'Fecha De Recepcion', 
        'Cant. Recibida', 'U/M', 'Valor', 'N° N/E'
    ]
    ws.append(headers)

    # Estilo Encabezado
    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 4. Iterar y Escribir
    for e in entradas:
        # Lógica condicional por tipo de ítem
        codigo = "-"
        descripcion = e.descripcion_entrada or "-"
        nro_parte = "-"
        um = "-"

        if e.material:
            codigo = e.material.codigo
            descripcion = descripcion if descripcion != "-" else e.material.descripcion
            nro_parte = e.material.nro_parte or "-"
            um = e.material.unidad_medida
        elif e.activo:
            codigo = e.activo.codigo_activo
            descripcion = descripcion if descripcion != "-" else e.activo.descripcion
            nro_parte = e.activo.nro_parte_ac or "-"
            um = e.activo.unidad_medida_ac
        elif e.gasto_directo:
            codigo = e.gasto_directo.codigo_dg
            descripcion = descripcion if descripcion != "-" else e.gasto_directo.descripcion
            nro_parte = e.gasto_directo.nro_parte_dg or "-"
            um = e.gasto_directo.unidad_medida_dg

        # Cálculos Financieros
        pu = float(e.precio_unitario or 0)
        total_solicitado = float(e.cantidad_solicitada) * pu
        valor_recibido = float(e.cantidad_recibida) * pu

        row = [
            e.nro_rq or "-",                                # No. Requisicion
            codigo,                                          # CODIGO MATERIAL
            descripcion,                                     # Descripción Del Material
            nro_parte,                                       # Numero de Parte
            e.departamento or "-",                           # Cargo
            e.nro_odc,                                       # No. Orden de Compra
            e.fecha_firma_odc.strftime('%d/%m/%Y') if e.fecha_firma_odc else "-", # Fecha Firma ODC
            e.proveedor,                                     # Proveedor
            e.moneda,                                        # Moneda
            float(e.cantidad_solicitada),                    # Cant. Solicitada
            um,                                              # U/M
            pu,                                              # U.P.
            total_solicitado,                                # Total
            e.eta.strftime('%d/%m/%Y') if e.eta else "-",    # ETA
            e.fecha_recepcion.strftime('%d/%m/%Y'),          # Fecha De Recepcion
            float(e.cantidad_recibida),                      # Cant. Recibida
            um,                                              # U/M
            valor_recibido,                                  # Valor
            e.nro_nota_entrega                               # N° N/E
        ]
        ws.append(row)

    # Ajuste de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    # 5. Respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Historial_Entradas_{datetime.date.today()}.xlsx'
    wb.save(response)
    return response


# ==========================================
# VISTA: IMPORTAR WHITE LIST (LISTA BLANCA)
# ==========================================
@login_required(login_url='login')
@user_passes_test(lambda u: u.is_superuser, login_url='dashboard')
def importar_whitelist(request):
    import re
    from django.contrib import messages

    if request.method == 'POST':
        correos_crudos = request.POST.get('correos_crudos', '').strip()
        
        if not correos_crudos:
            messages.warning(request, "Por favor, ingrese al menos un correo electrónico.")
            return redirect('importar_whitelist')

        # Extraer correos usando split o expresiones regulares
        # Separadores permitidos: comas, espacios, saltos de línea, punto y coma
        raw_list = re.split(r'[\s,;\n\r]+', correos_crudos)
        emails = []
        email_regex = re.compile(r'^[\w\.-]+@[\w\.-]+\.\w+$')

        for raw in raw_list:
            clean_email = raw.strip().lower()
            if clean_email and email_regex.match(clean_email):
                emails.append(clean_email)

        # Eliminar duplicados en el lote enviado
        emails_unicos = list(set(emails))

        if not emails_unicos:
            messages.error(request, "No se encontraron correos con formato válido.")
            return redirect('importar_whitelist')

        try:
            with transaction.atomic():
                cant_antes = CorreoAutorizado.objects.count()
                
                # Crear objetos para bulk_create
                to_create = [CorreoAutorizado(email=email) for email in emails_unicos]
                CorreoAutorizado.objects.bulk_create(to_create, ignore_conflicts=True)
                
                cant_despues = CorreoAutorizado.objects.count()
                diferencia = cant_despues - cant_antes

            messages.success(
                request, 
                f"Proceso completado. Se importaron {diferencia} correos nuevos exitosamente a la Lista Blanca (de {len(emails_unicos)} válidos procesados)."
            )
            return redirect('importar_whitelist')
            
        except Exception as e:
            messages.error(request, f"Error al realizar la importación masiva: {str(e)}")
            return redirect('importar_whitelist')

    return render(request, 'inventario/importar_whitelist.html')


# VISTA: Gestión de Equipo (Panel de Control de Usuarios para Administradores)
@login_required(login_url='login')
@user_passes_test(lambda u: u.is_superuser, login_url='dashboard')
def gestion_equipo(request):
    from django.contrib.auth.models import User, Group
    from django.contrib import messages
    from django.db import transaction

    # Definir los roles permitidos por el sistema WMS
    grupos_permitidos = ['Coordinador de Almacén', 'Especialista de Activos', 'Analista de Almacén', 'Gerente de Procura']

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'crear':
            cedula = request.POST.get('cedula', '').strip()
            nombre = request.POST.get('nombre', '').strip().upper()
            password = request.POST.get('password', '').strip()
            grupo_nombre = request.POST.get('rol', '').strip()

            if not cedula or not nombre or not password or not grupo_nombre:
                messages.error(request, "Error: Todos los campos son obligatorios para registrar un nuevo integrante.")
            elif User.objects.filter(username=cedula).exists():
                messages.error(request, f"Error: Ya existe un usuario registrado con la Cédula '{cedula}'.")
            elif grupo_nombre not in grupos_permitidos:
                messages.error(request, "Error: El rol seleccionado no es válido en el sistema.")
            else:
                try:
                    with transaction.atomic():
                        # Crear el usuario en el modelo User
                        user = User.objects.create(
                            username=cedula,
                            first_name=nombre,
                            is_active=True
                        )
                        user.set_password(password)

                        # Mapear permisos según el rol
                        # Coordinador, Especialista y Analista tienen acceso operacional (is_staff = True)
                        if grupo_nombre in ['Coordinador de Almacén', 'Especialista de Activos', 'Analista de Almacén']:
                            user.is_staff = True

                        if grupo_nombre == 'Coordinador de Almacén':
                            user.is_superuser = True

                        user.save()

                        # Crear/Buscar el grupo en Django y asignarlo
                        grupo_obj, _ = Group.objects.get_or_create(name=grupo_nombre)
                        user.groups.add(grupo_obj)

                        messages.success(request, f"Éxito: Usuario con Cédula '{cedula}' ({nombre}) creado correctamente con el rol de '{grupo_nombre}'.")
                except Exception as e:
                    messages.error(request, f"Error inesperado al registrar el usuario: {str(e)}")
            return redirect('gestion_equipo')

        elif action == 'toggle_status':
            user_id = request.POST.get('user_id')
            if user_id:
                try:
                    user_to_toggle = User.objects.get(id=user_id)
                    # Protección: No desactivarse a sí mismo
                    if user_to_toggle == request.user:
                        messages.error(request, "Seguridad: No puedes desactivar tu propia cuenta activa.")
                    else:
                        user_to_toggle.is_active = not user_to_toggle.is_active
                        user_to_toggle.save()
                        estado = "activado" if user_to_toggle.is_active else "desactivado"
                        messages.success(request, f"Éxito: El usuario {user_to_toggle.first_name} (Cédula: {user_to_toggle.username}) ha sido {estado} correctamente.")
                except User.DoesNotExist:
                    messages.error(request, "Error: El usuario especificado no existe.")
                except Exception as e:
                    messages.error(request, f"Error inesperado al cambiar estado: {str(e)}")
            return redirect('gestion_equipo')

        elif action == 'editar':
            user_id = request.POST.get('user_id')
            cedula = request.POST.get('cedula', '').strip()
            nombre = request.POST.get('nombre', '').strip().upper()
            grupo_nombre = request.POST.get('rol', '').strip()
            password = request.POST.get('password', '').strip()

            if not user_id or not cedula or not nombre or not grupo_nombre:
                messages.error(request, "Error: Todos los campos obligatorios deben completarse para actualizar al integrante.")
            elif grupo_nombre not in grupos_permitidos:
                messages.error(request, "Error: El rol seleccionado no es válido en el sistema.")
            else:
                try:
                    with transaction.atomic():
                        user_to_edit = User.objects.get(id=user_id)
                        
                        # Validar si cambió la cédula y si la nueva ya existe en otro usuario
                        if user_to_edit.username != cedula and User.objects.filter(username=cedula).exclude(id=user_id).exists():
                            raise ValueError(f"Ya existe otro usuario registrado con la Cédula '{cedula}'.")

                        # Actualizar datos básicos
                        user_to_edit.username = cedula
                        user_to_edit.first_name = nombre

                        # Si se ingresó una contraseña nueva, actualizarla
                        if password:
                            user_to_edit.set_password(password)

                        # Actualizar flags de staff y superusuario según el nuevo rol
                        user_to_edit.is_staff = False
                        user_to_edit.is_superuser = False

                        if grupo_nombre in ['Coordinador de Almacén', 'Especialista de Activos', 'Analista de Almacén']:
                            user_to_edit.is_staff = True

                        if grupo_nombre == 'Coordinador de Almacén':
                            user_to_edit.is_superuser = True

                        user_to_edit.save()

                        # Actualizar grupos/roles
                        user_to_edit.groups.clear()
                        grupo_obj, _ = Group.objects.get_or_create(name=grupo_nombre)
                        user_to_edit.groups.add(grupo_obj)

                        messages.success(request, f"Éxito: Datos del usuario '{cedula}' ({nombre}) actualizados correctamente.")
                except Exception as e:
                    messages.error(request, f"Error inesperado al editar el usuario: {str(e)}")
            return redirect('gestion_equipo')

    # GET: Listado de todos los usuarios
    usuarios = User.objects.all().order_by('-date_joined')
    
    contexto = {
        'usuarios': usuarios,
        'grupos_permitidos': grupos_permitidos,
    }
    return render(request, 'inventario/gestion_equipo.html', contexto)


# VISTA: Cambio de Contraseña Obligatorio (Primer Inicio de Sesión)
from django.contrib.auth.views import PasswordChangeView
from django.urls import reverse_lazy
from django.contrib import messages

class CustomPasswordChangeView(PasswordChangeView):
    template_name = 'inventario/password_change_form.html'
    success_url = reverse_lazy('dashboard')

    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Desactivar la bandera de cambio obligatorio en el perfil del usuario
        from inventario.models import PerfilUsuario
        perfil, _ = PerfilUsuario.objects.get_or_create(user=self.request.user)
        perfil.debe_cambiar_clave = False
        perfil.save()
        
        messages.success(self.request, "Éxito: Tu contraseña ha sido actualizada correctamente. ¡Bienvenido al sistema WMS!")
        return response


# VISTA: Perfil de Usuario (Visualización y Edición de Foto/Datos)
@login_required
def perfil_usuario(request):
    from inventario.models import PerfilUsuario
    from inventario.forms import UserUpdateForm, PerfilUpdateForm

    # Obtener o crear el perfil asociado al usuario
    perfil, _ = PerfilUsuario.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        u_form = UserUpdateForm(request.POST, instance=request.user)
        p_form = PerfilUpdateForm(request.POST, request.FILES, instance=perfil)

        if u_form.is_valid() and p_form.is_valid():
            try:
                with transaction.atomic():
                    u_form.save()
                    p_form.save()
                messages.success(request, "Éxito: Tu información de perfil ha sido actualizada correctamente.")
                return redirect('perfil_usuario')
            except Exception as e:
                messages.error(request, f"Error inesperado al guardar el perfil: {str(e)}")
        else:
            messages.error(request, "Error: Por favor verifica los datos ingresados en el formulario.")
    else:
        u_form = UserUpdateForm(instance=request.user)
        p_form = PerfilUpdateForm(instance=perfil)

    contexto = {
        'u_form': u_form,
        'p_form': p_form,
        'perfil': perfil,
    }
    return render(request, 'inventario/perfil.html', contexto)
